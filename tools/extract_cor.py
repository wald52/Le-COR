#!/usr/bin/env python3
"""Extraction reproductible des séries du COR depuis les fichiers Excel.

Lit les fichiers Excel officiels rangés sous `data/Données du COR/` et génère
deux fichiers consommés par le site :
  - `data/cor-series.generated.js`   (window.COR_SERIES)  — graphiques essentiels
  - `data/cor-explorer.generated.js` (window.COR_EXPLORER) — données de l'explorateur
L'explorateur, de loin le plus gros bloc, est séparé pour être chargé
paresseusement côté client (chargement initial allégé).

But : remplacer les valeurs d'amorçage par les CHIFFRES OFFICIELS, de façon
traçable (réexécuter ce script régénère les données).

Lancement :  python3 tools/extract_cor.py
Dépendance :  openpyxl
"""
import glob, os, json, types, openpyxl
from openpyxl.worksheet import print_settings

# Certains classeurs du COR (ex. partie 3 de juin 2026) contiennent une zone
# d'impression invalide (« #N/A ») qu'openpyxl refuse : on la tolère.
_orig_print_titles = print_settings.PrintTitles.from_string.__func__


def _safe_print_titles(cls, value):
    try:
        return _orig_print_titles(cls, value)
    except Exception:
        return types.SimpleNamespace(rows=None, cols=None)


print_settings.PrintTitles.from_string = classmethod(_safe_print_titles)


def write_js(dest, varname, obj):
    """Émet `window.<varname> = JSON.parse("…")` minifié.

    Le moteur JS analyse cette forme nettement plus vite qu'un littéral objet
    équivalent (chemin JSON optimisé de V8), et le fichier est ~3× plus léger
    qu'en indenté — ce qui réduit d'autant le temps de parsing au chargement
    (Total Blocking Time). La double sérialisation produit un littéral de chaîne
    JS sûr.
    """
    compact = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    with open(dest, "w", encoding="utf-8") as f:
        f.write("/* FICHIER GÉNÉRÉ — ne pas éditer à la main.\n")
        f.write("   Source : fichiers Excel officiels du COR (data/Données du COR/).\n")
        f.write("   Régénérer avec : python3 tools/extract_cor.py */\n")
        f.write("window.%s = JSON.parse(" % varname)
        f.write(json.dumps(compact, ensure_ascii=False))
        f.write(");\n")


BASE = os.path.join(os.path.dirname(__file__), "..", "data", "Données du COR")

# Millésime -> (préfixe dossier, motif de fichier, productivité de référence du
# scénario à tracer). Pour 2024/2025, la synthèse ne contient que le scénario
# de référence (ligne « Sc. Ref »), d'où prod_ref = None.
VINTAGES = [
    ("2016", "2016-06", "Indicateurs financiers", 0.013),
    ("2017", "2017-06", "Indicateurs financiers", 0.013),
    ("2018", "2018-06", "indicateurs financiers", 0.013),
    ("2019", "2019-06", "Partie 2", 0.013),
    ("2020", "2020-11", "Partie 2", 0.013),
    ("2021", "2021-06", "Partie 2", 0.013),
    ("2022", "2022-09", "septembre 2022 - partie 2", 0.013),
    ("2023", "2023-06", "synthèse", 0.010),
    ("2024", "2024-06", "synthèse", None),
    ("2025", "2025-06", "synthèse", None),
    ("2026", "2026-06", "synthèse", None),
]

# Millésime le plus récent : sert de série « réalisé » et de base à
# l'explorateur d'indicateurs.
LATEST = "2026"

# Productivité de référence affichée dans la légende (en %).
PROD_LABEL = {"2016": "1,3", "2017": "1,3", "2018": "1,3", "2019": "1,3",
              "2020": "1,3", "2021": "1,3", "2022": "1,3", "2023": "1,0",
              "2024": "1,0", "2025": "0,7", "2026": "0,7"}

# Couleur par millésime : dégradé gris → bleu → vert → orange → rouge,
# pour lire visuellement l'écoulement du temps (2026 = le plus saillant).
COLORS = {"2016": "#9aa7b4", "2017": "#7d8ca0", "2018": "#5b6f93",
          "2019": "#3f7cb0", "2020": "#2ca089", "2021": "#6aa84f",
          "2022": "#e0a800", "2023": "#e8731c", "2024": "#d6452a",
          "2025": "#c2185b", "2026": "#7b1fa2"}


def find_depenses_block(wb):
    """Localise le bloc de données « Dépenses, en % du PIB » / ligne 'Obs'."""
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for i, r in enumerate(rows):
            if len(r) > 2 and r[1] == "Dépenses, en % du PIB" and r[2] == "Obs":
                return rows, i
    return None, None


def year_cols(rows):
    for r in rows[:8]:
        if any(isinstance(c, int) and 1990 <= c <= 2100 for c in r):
            return {i: c for i, c in enumerate(r)
                    if isinstance(c, int) and 1990 <= c <= 2100}
    return {}


def to_series(row, ycols):
    """Convertit une ligne en {année: valeur en %} (les parts sont en fraction)."""
    return {ycols[i]: round(row[i] * 100, 3)
            for i in ycols if i < len(row) and isinstance(row[i], (int, float))}


def extract_depenses(path, prod_ref):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows, iobs = find_depenses_block(wb)
    if rows is None:
        wb.close()
        return None
    ycols = year_cols(rows)
    observed = to_series(rows[iobs], ycols)
    projection = {}
    # La projection de référence est dans l'une des ~7 lignes suivant l'observé.
    for r in rows[iobs + 1:iobs + 8]:
        c2 = r[2] if len(r) > 2 else None
        if c2 == "Sc. Ref":
            projection = to_series(r, ycols)
            break
        if prod_ref is not None and isinstance(c2, (int, float)) \
                and abs(float(c2) - prod_ref) < 1e-4:
            projection = to_series(r, ycols)
            break
    wb.close()
    return {"observed": observed, "projection": projection}


def first_file(folder_prefix, name_substr):
    folders = glob.glob(os.path.join(BASE, folder_prefix + "*"))
    if not folders:
        return None
    for p in glob.glob(os.path.join(folders[0], "*.xlsx")):
        if name_substr.lower() in os.path.basename(p).lower():
            return p
    return None


# --------------------------------------------------------------------------
# Solde / Ressources — multi-millésimes (2016 → 2026)
#
# Chaque rapport publie le solde et les ressources « observés et projetés »,
# mais la mise en page varie selon l'époque :
#   - mode "sdr"   (2023+)      : feuille de synthèse, lignes Dépenses /
#                                 Ressources / Solde (convention EPR) ;
#   - mode "conv"  (2020-2022)  : figure avec une ligne « Observé » puis un
#                                 bloc par convention comptable (EEC/TCC/EPR),
#                                 une ligne par scénario de productivité ;
#   - mode "title" (2016-2019)  : figure avec un bloc « libellé | Obs » puis
#                                 une ligne par scénario (convention « COR »).
# On retient le scénario de productivité du tableau VINTAGES (1,3 % avant
# 2023) et la convention EPR dès qu'elle existe, « COR » avant 2020.
# --------------------------------------------------------------------------
SOLDE_SOURCES = [
    ("2016", "2016-06", "Indicateurs financiers", "title", ("solde financier", "projeté"), 0.013, "conv. COR"),
    ("2017", "2017-06", "Indicateurs financiers", "title", ("solde financier", "projeté"), 0.013, "conv. COR"),
    ("2018", "2018-06", "indicateurs financiers", "title", ("solde financier", "projeté"), 0.013, "conv. COR"),
    ("2019", "2019-06", "Partie 2", "title", ("solde financier annuel observé",), 0.013, "conv. COR"),
    ("2020", "2020-11", "Partie 2", "conv", ("solde financier observé et projeté",), 0.013, "conv. EPR"),
    ("2021", "2021-06", "Partie 2", "conv", ("solde observé et projeté",), 0.013, "conv. EPR"),
    ("2022", "2022-09", "septembre 2022 - partie 2", "conv", ("solde observé et projeté",), 0.013, "conv. EPR"),
    ("2023", "2023-06", "synthèse", "sdr", ("Solde_dép_ress", "Solde"), None, "conv. EPR"),
    ("2024", "2024-06", "synthèse", "sdr", ("Solde dépenses ressources", "Solde"), None, "conv. EPR"),
    ("2025", "2025-06", "synthèse", "sdr", ("Solde dépenses ressources", "Solde"), None, "conv. EPR"),
    ("2026", "2026-06", "synthèse", "sdr", ("Solde dépenses ressources", "Solde"), None, "conv. EPR"),
]
RESSOURCES_SOURCES = [
    ("2016", "2016-06", "Indicateurs financiers", "title", ("ressources", "dépenses du système"), 0.013, "conv. COR"),
    ("2017", "2017-06", "Indicateurs financiers", "title", ("ressources", "projet"), 0.013, "conv. COR"),
    ("2018", "2018-06", "indicateurs financiers", "title", ("ressources", "observ"), 0.013, "conv. COR"),
    ("2019", "2019-06", "Partie 2", "title", ("ressources observées", "projet"), 0.013, "conv. COR"),
    ("2020", "2020-11", "Partie 2", "conv", ("ressources observées et projetées",), 0.013, "conv. EPR"),
    ("2021", "2021-06", "Partie 2", "conv", ("ressources observées et projetées",), 0.013, "conv. EPR"),
    ("2022", "2022-09", "septembre 2022 - partie 2", "conv", ("ressources observées et projetées",), 0.013, "conv. EPR"),
    ("2023", "2023-06", "synthèse", "sdr", ("Solde_dép_ress", "Ressources"), None, "conv. EPR"),
    ("2024", "2024-06", "synthèse", "sdr", ("Solde dépenses ressources", "Ressources"), None, "conv. EPR"),
    ("2025", "2025-06", "synthèse", "sdr", ("Solde dépenses ressources", "Ressources"), None, "conv. EPR"),
    ("2026", "2026-06", "synthèse", "sdr", ("Solde dépenses ressources", "Ressources"), None, "conv. EPR"),
]


def extract_sdr(path, sheet):
    """Renvoie {ligne: {année: % PIB}} pour Dépenses / Ressources / Solde."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    ycols = {}
    for r in rows[:6]:
        ints = {i: c for i, c in enumerate(r)
                if isinstance(c, int) and 1990 <= c <= 2100}
        if ints:
            ycols = ints
            break
    out = {}
    for r in rows:
        lbl = r[1] if len(r) > 1 else None
        if lbl in ("Dépenses", "Ressources", "Solde"):
            out[lbl] = {ycols[i]: round(r[i] * 100, 3)
                        for i in ycols if i < len(r) and isinstance(r[i], (int, float))}
    wb.close()
    return out


def _norm(s):
    # Apostrophe typographique → droite : les titres du COR mêlent « d'emploi »
    # et « d'emploi » d'une édition à l'autre, ce qui cassait l'appariement.
    return " ".join(str(s).replace("’", "'").split()).lower()


def sheet_by_title(wb, keys, exclude=()):
    """Première feuille dont le titre (A1) contient tous les mots-clés `keys`
    et aucun des mots-clés `exclude`."""
    for ws in wb.worksheets:
        a1 = _norm(ws.cell(1, 1).value or "")
        if all(k in a1 for k in keys) and not any(x in a1 for x in exclude):
            return ws
    return None


def _sheet_ycols(rows):
    for r in rows[:8]:
        ints = {i: c for i, c in enumerate(r)
                if isinstance(c, int) and 1990 <= c <= 2100}
        if len(ints) >= 3:
            return ints
    return {}


def extract_title_block(path, keys, prod_ref):
    """Mode 2016-2019 : bloc « libellé | Obs » puis lignes-scénarios (col C)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, keys)
    if ws is None:
        wb.close()
        return None
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ycols = _sheet_ycols(rows)
    iobs = next((i for i, r in enumerate(rows)
                 if len(r) > 2 and str(r[2]).strip() == "Obs"), None)
    if iobs is None or not ycols:
        return None
    for r in rows[iobs + 1:iobs + 8]:
        c2 = r[2] if len(r) > 2 else None
        if isinstance(c2, (int, float)) and abs(float(c2) - prod_ref) < 1e-4:
            return to_series(r, ycols)
    return None


def extract_conv_block(path, keys, prod_ref, convention="EPR"):
    """Mode 2020-2022 : ligne « Observé » puis un bloc par convention comptable."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, keys)
    if ws is None:
        wb.close()
        return None
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ycols = _sheet_ycols(rows)
    if not ycols:
        return None
    iconv = next((i for i, r in enumerate(rows) if len(r) > 1 and r[1]
                  and _norm(r[1]) == _norm(f"Convention {convention}")), None)
    if iconv is None:
        return None
    for i in range(iconv, min(iconv + 6, len(rows))):
        r = rows[i]
        c1 = r[1] if len(r) > 1 else None
        if i > iconv and isinstance(c1, str) and c1.strip().lower().startswith("convention"):
            break
        c2 = r[2] if len(r) > 2 else None
        if isinstance(c2, (int, float)) and abs(float(c2) - prod_ref) < 1e-4:
            return to_series(r, ycols)
    return None


def extract_fin_multi(sources, what):
    """Projections multi-millésimes {vy: {année: % PIB}} pour solde/ressources."""
    out = {}
    for vy, dpat, fpat, mode, keys, prod, conv in sources:
        path = first_file(dpat, fpat)
        if not path:
            print(f"✗ {what} {vy} : fichier introuvable")
            continue
        if mode == "sdr":
            sdr = extract_sdr(path, keys[0])
            serie = sdr.get(keys[1]) if sdr else None
        elif mode == "conv":
            serie = extract_conv_block(path, keys, prod)
        else:
            serie = extract_title_block(path, keys, prod)
        if serie:
            out[vy] = serie
        else:
            print(f"✗ {what} {vy} : bloc {mode} introuvable {keys}")
    return out


# --------------------------------------------------------------------------
# Niveau de vie relatif et âge conjoncturel — multi-millésimes (2023 → 2026)
# (les synthèses antérieures à 2023 ne publient pas ces feuilles)
# --------------------------------------------------------------------------
NV_SOURCES = [
    ("2023", "2023-06", "Niveau_vie", 0.010),
    ("2024", "2024-06", "Niveau de vie relatif", "Scénario de référence"),
    ("2025", "2025-06", "Niveau de vie relatif", "Scénario de référence"),
    ("2026", "2026-06", "Niveau de vie relatif", "Scénario de référence"),
]
AGE_SOURCES = [
    ("2023", "2023-06", "Âge_retraite"),
    ("2024", "2024-06", "Âge conjoncturel"),
    ("2025", "2025-06", "Âge conjoncturel"),
    ("2026", "2026-06", "Âge conjoncturel"),
]


def extract_nv_vintage(path, sheet, selector):
    """Niveau de vie relatif : (observé, scénario de référence) en %.

    `selector` : libellé « Scénario de référence » (2024+) ou productivité de
    référence en fraction (2023, où les projections sont des lignes 1,6/1,3/1,0 %).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None, None
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    ycols = {}
    for r in rows[:6]:
        ints = {i: c for i, c in enumerate(r) if isinstance(c, int) and 1960 <= c <= 2100}
        if ints:
            ycols = ints
            break
    obs = ref = None
    for r in rows:
        c2 = r[2] if len(r) > 2 else None
        vals = {ycols[i]: round(r[i] * 100, 1)
                for i in ycols if i < len(r) and isinstance(r[i], (int, float))}
        if c2 == "Observations":
            obs = vals
        elif isinstance(selector, str) and c2 == selector:
            ref = vals
        elif isinstance(selector, float) and isinstance(c2, (int, float)) \
                and abs(float(c2) - selector) < 1e-4:
            ref = vals
    return obs, ref


def extract_age_vintage(path, sheet):
    """Âge conjoncturel : (observé, projeté « Tous scénarios »), en années."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cand = [s for s in wb.sheetnames if s.strip() == sheet.strip()]
    if not cand:
        wb.close()
        return None, None
    rows = list(wb[cand[0]].iter_rows(values_only=True))
    wb.close()
    return _obs_proj(rows, 1)


def labeled_row(path, sheet_keys, row_key, scale=1.0):
    """Série {année: valeur} de la 1re ligne dont le libellé contient row_key.

    Le libellé peut être en colonne A (rapports ≤ 2019) ou B (rapports
    récents) ; les années sont repérées par le premier bloc d'entiers.
    """
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, sheet_keys)
    if ws is None:
        wb.close()
        return {}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ym = _ymap(rows)
    for r in rows:
        for j in (0, 1):
            v = r[j] if len(r) > j else None
            if isinstance(v, str) and row_key.lower() in _norm(v):
                serie = {ym[i]: round(r[i] * scale, 3) for i in ym
                         if i < len(r) and isinstance(r[i], (int, float))}
                if serie:  # ignore les faux positifs sans données (ex. titre A1)
                    return serie
    return {}


def labeled_row_any(path, sheet_keys, row_keys, scale=1.0):
    """Comme labeled_row, mais essaie plusieurs clés et renvoie (série, libellé).

    Le libellé exact de la ligne retenue permet d'annoter la légende (il porte
    souvent la valeur de l'hypothèse : « Tous scénarios 7% », « Scénario de
    référence 0,7% »…).
    """
    if not path:
        return {}, None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, sheet_keys)
    if ws is None:
        wb.close()
        return {}, None
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ym = _ymap(rows)
    for key in row_keys:
        for r in rows:
            for j in (0, 1):
                v = r[j] if len(r) > j else None
                if isinstance(v, str) and key.lower() in _norm(v):
                    serie = {ym[i]: round(r[i] * scale, 3) for i in ym
                             if i < len(r) and isinstance(r[i], (int, float))}
                    if serie:
                        return serie, v.strip()
    return {}, None


def block_labeled_row(path, sheet_keys, block_key, row_key, scale=1.0):
    """Ligne `row_key` à l'intérieur du bloc `block_key` d'une feuille.

    Utile quand une feuille empile plusieurs blocs au même format (ex. les
    rapports démographiques 20-59/60+ puis 20-64/65+).
    """
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, sheet_keys)
    if ws is None:
        wb.close()
        return {}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ym = _ymap(rows)
    start = None
    for i, r in enumerate(rows):
        for j in (0, 1):
            v = r[j] if len(r) > j else None
            if isinstance(v, str) and _norm(block_key) in _norm(v):
                start = i
                break
        if start is not None:
            break
    if start is None:
        return {}
    for r in rows[start:start + 10]:
        for j in (0, 1):
            v = r[j] if len(r) > j else None
            if isinstance(v, str) and row_key.lower() in _norm(v):
                return {ym[i]: round(r[i] * scale, 3) for i in ym
                        if i < len(r) and isinstance(r[i], (int, float))}
    return {}


def _hyp_note(label):
    """Extrait la valeur d'hypothèse d'un libellé (« Tous scénarios 7% » → « 7 % »)."""
    import re
    m = re.search(r"([\d]+(?:,\d+)?\s*%)", label or "")
    return m.group(1).replace("%", " %").replace("  ", " ") if m else None


def _pct_fix(serie, ceiling=120):
    """Corrige l'échelle quand un fichier stocke déjà des % et non des fractions.

    Exemple : le taux de chômage du rapport 2017 est saisi « 8,5 » là où les
    autres millésimes écrivent « 0,085 » ; multiplié par 100 il donnerait 850 %.
    """
    if serie and max(serie.values()) > ceiling:
        return {y: round(v / 100, 3) for y, v in serie.items()}
    return serie


def _block_sub_ref(rows, label_key, ref):
    """Sous-ligne du bloc `label_key` : col C == `ref` (texte) ou ≈ ref (taux)."""
    ym = _ymap(rows)
    start = None
    for i, r in enumerate(rows):
        if len(r) > 1 and r[1] and label_key.lower() in str(r[1]).lower():
            start = i
            break
    if start is None:
        return {}
    for r in rows[start:start + 8]:
        c2 = r[2] if len(r) > 2 else None
        if (isinstance(ref, str) and str(c2).strip() == ref) or \
                (isinstance(ref, float) and isinstance(c2, (int, float))
                 and abs(float(c2) - ref) < 1e-4):
            return {ym[i]: r[i] for i in ym if i < len(r) and isinstance(r[i], (int, float))}
    return {}


def pension_ratio_vintage(path, ref):
    """Pension moyenne / rémunération moyenne (%), scénario de référence.

    `ref` : « Sc. Ref » (2024+) ou productivité de référence en fraction
    (2023, où les scénarios sont des lignes 1,6/1,3/1,0/0,7 %).
    """
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(max_row=14, max_col=3, values_only=True))
        labels = [str(r[1]).lower() for r in rows if len(r) > 1 and r[1]]
        if any("pension moyenne" in l for l in labels) \
                and any("rémunération nett" in l for l in labels):
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            pen = _block_sub_ref(rows, "pension moyenne", ref)
            rem = _block_sub_ref(rows, "rémunération nett", ref)
            return _ratio(pen, rem, 100)
    wb.close()
    return {}


def extract_fecondite(path):
    """Indice de fécondité : observé (définitif + provisoire) et scénario central."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = None
    for s in wb.worksheets:
        a1 = str(s.cell(1, 1).value or "").lower()
        if "fécondit" in a1 and ("observé" in a1 or "projet" in a1):
            ws = s
            break
    if ws is None:
        wb.close()
        return None
    rows = list(ws.iter_rows(values_only=True))
    ycols = {}
    for r in rows[:6]:
        ic = {i: c for i, c in enumerate(r) if isinstance(c, int) and 1990 <= c <= 2100}
        if ic:
            ycols = ic
            break
    central, obs = {}, {}
    for r in rows:
        lbl = str(r[1]).lower() if len(r) > 1 and r[1] else ""
        vals = {ycols[i]: round(r[i], 3) for i in ycols
                if i < len(r) and isinstance(r[i], (int, float))}
        if "central" in lbl and not central:
            central = vals
        if lbl.startswith("observé") or "données provisoires" in lbl:
            obs.update(vals)
    wb.close()
    return {"observed": obs, "central": central}


def extract_productivite_obs(path):
    """Croissance annuelle observée de la productivité (Fig 1.9 / 1.10)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = None
    for s in wb.worksheets:
        a1 = str(s.cell(1, 1).value or "").lower()
        if "productivit" in a1 and "croissance" in a1:
            ws = s
            break
    if ws is None:
        wb.close()
        return None
    rows = list(ws.iter_rows(values_only=True))
    ycols = {}
    for r in rows[:6]:
        ic = {i: c for i, c in enumerate(r) if isinstance(c, int) and 1980 <= c <= 2100}
        if ic:
            ycols = ic
            break
    annual = {}
    for r in rows:
        lbl = str(r[1]).lower() if len(r) > 1 and r[1] else ""
        if lbl.startswith("croissance annuelle observ"):
            annual = {ycols[i]: r[i] * 100 for i in ycols
                      if i < len(r) and isinstance(r[i], (int, float))}
            break
    wb.close()
    return annual


def moving_average(series, window=5):
    """Moyenne mobile centrée d'un dict {année: valeur}."""
    ys = sorted(series)
    out = []
    half = window // 2
    for y in ys:
        vals = [series[k] for k in range(y - half, y + half + 1) if k in series]
        if len(vals) >= 3:
            out.append({"x": y, "y": round(sum(vals) / len(vals), 3)})
    return out


# ==========================================================================
# EXPLORATEUR D'INDICATEURS — catalogue de séries « observé + projeté »
# (un seul graphique sur le site, indicateur sélectionnable). Permet d'avoir
# « tous les indicateurs » sans empiler les graphiques.
# ==========================================================================
R26 = "2026-06"  # on s'appuie sur le rapport le plus récent
R25 = "2025-06"  # rapport précédent (certains fichiers n'existent qu'en 2025)


def _rows(filepat, sheet, vintage=R26):
    path = first_file(vintage, filepat)
    if not path:
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cand = [s for s in wb.sheetnames if s.strip() == sheet.strip()]
    if not cand:
        wb.close()
        return None
    rows = list(wb[cand[0]].iter_rows(values_only=True))
    wb.close()
    return rows


def _ymap(rows):
    """Premier bloc contigu d'années croissantes (gère les blocs répétés côte à côte)."""
    for r in rows[:6]:
        ic = [(i, c) for i, c in enumerate(r) if isinstance(c, int) and 1950 <= c <= 2100]
        if len(ic) >= 3:
            out, last = {}, None
            for i, c in ic:
                if last is not None and c < last:
                    break
                out[i] = c
                last = c
            return out
    return {}


def _row_label(rows, key, scale=1.0):
    """Première ligne dont col1 contient `key`, mappée aux années (1er bloc)."""
    ym = _ymap(rows)
    for r in rows:
        if len(r) > 1 and r[1] and key.lower() in str(r[1]).lower():
            return {ym[i]: round(r[i] * scale, 3) for i in ym
                    if i < len(r) and isinstance(r[i], (int, float))}
    return {}


# --------------------------------------------------------------------------
# Indicateurs « par génération » (parties 3 & 4) — superposition multi-rapports.
#
# Ces figures ont un en-tête d'années/générations horizontal, mais : (a) les
# cohortes descendent sous 1950 (jusqu'à 1906), hors plancher de `_ymap` ; (b)
# l'en-tête est parfois stocké en texte ('1906') ; (c) le numéro de figure
# change d'un rapport à l'autre (repérage par titre, pas par feuille). D'où des
# helpers dédiés, sans toucher à ceux des parties 1 & 2.
# --------------------------------------------------------------------------
def _xmap(rows, floor=1900):
    """Comme `_ymap`, mais plancher paramétrable et en-têtes d'année en texte."""
    for r in rows[:8]:
        ic = []
        for i, c in enumerate(r):
            y = None
            if isinstance(c, int) and floor <= c <= 2100:
                y = c
            elif isinstance(c, str) and c.strip().isdigit() and floor <= int(c) <= 2100:
                y = int(c)
            if y is not None:
                ic.append((i, y))
        if len(ic) >= 3:
            out, last = {}, None
            for i, c in ic:
                if last is not None and c < last:
                    break
                out[i] = c
                last = c
            return out
    return {}


def _plausible(serie, lo, hi):
    """Garde la série seulement si sa médiane tombe dans [lo, hi], sinon {}.

    Filet de sécurité quand le titre d'une figure matche, dans un vieux rapport,
    une variante hors sujet (ex. « écart d'âge » au lieu de « âge moyen »)."""
    if not serie:
        return {}
    v = sorted(serie.values())
    return serie if lo <= v[len(v) // 2] <= hi else {}


def file_with_sheet(vpfx, sheet_keys, exclude=()):
    """Premier .xlsx du dossier `vpfx*` contenant une feuille au titre matchant."""
    folders = glob.glob(os.path.join(BASE, vpfx + "*"))
    if not folders:
        return None
    for p in sorted(glob.glob(os.path.join(folders[0], "*.xlsx"))):
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception:
            continue
        hit = sheet_by_title(wb, sheet_keys, exclude) is not None
        wb.close()
        if hit:
            return p
    return None


def pick_cohort_row(path, sheet_keys, row_keys=(), scale=1.0, floor=1900,
                    prod_ref=None, block_key=None, exclude=()):
    """Série {x: val} d'une figure à en-tête d'années/générations horizontal.

    Retient la 1re ligne selon, dans l'ordre : une des `row_keys` (texte, col 0-2) ;
    un nombre ≈ `prod_ref` (col 1-2, sélection du scénario de référence dans les
    rapports qui empilent les scénarios de productivité) ; à défaut la 1re ligne
    portant ≥5 points qui ne soit pas la ligne d'en-tête. `block_key` borne la
    recherche au bloc suivant ce libellé."""
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, sheet_keys, exclude)
    if ws is None:
        wb.close()
        return {}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ym = _xmap(rows, floor)
    if not ym:
        return {}
    start = 0
    if block_key:
        start = next((i for i, r in enumerate(rows)
                      if any(isinstance(r[j], str) and _norm(block_key) in _norm(r[j])
                             for j in range(min(3, len(r))))), None)
        if start is None:
            return {}
    scope = rows[start:]

    def extract(r):
        return {ym[i]: round(r[i] * scale, 3) for i in ym
                if i < len(r) and isinstance(r[i], (int, float))}

    def ok(s):  # écarte la ligne d'en-tête (valeurs ≈ années) et les lignes courtes
        return len(s) >= 5 and not all(abs(v - x) < 0.5 for x, v in s.items())

    for key in row_keys:
        if not key:
            continue
        for r in scope:
            for j in (0, 1, 2):
                v = r[j] if len(r) > j else None
                if isinstance(v, str) and key.lower() in _norm(v):
                    s = extract(r)
                    if ok(s):
                        return s
    if prod_ref is not None:
        for r in scope:
            for j in (1, 2):
                v = r[j] if len(r) > j else None
                if isinstance(v, (int, float)) and abs(float(v) - prod_ref) < 1e-4:
                    s = extract(r)
                    if ok(s):
                        return s
    for r in scope:
        s = extract(r)
        if ok(s):
            return s
    return {}


def vertical_ratio(path, sheet_keys, scale=100, floor=1990, hi=2070):
    """Série {année: ratio×scale} d'une figure en colonnes (une année par ligne).

    Pour l'écart de pension F/H (Fig 3.24) : on retient la *dernière* valeur
    fractionnaire (0<v<2) de la ligne après l'année. C'est la colonne la plus à
    droite — le ratio « pension totale y compris majorations », renseigné en
    observé puis en projeté, donc cohérent sur tout l'horizon (contrairement à la
    1re colonne, « droit direct », qui s'arrête à l'observé)."""
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, sheet_keys)
    if ws is None:
        wb.close()
        return {}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for r in rows:
        yr = val = None
        for c in r:
            if yr is None and isinstance(c, int) and floor <= c <= hi:
                yr = c
            elif yr is not None and isinstance(c, (int, float)) and 0 < c < 2:
                val = c  # garde la dernière (colonne la plus à droite)
        if yr is not None and val is not None:
            out[yr] = round(val * scale, 3)
    return out


def overlay_vintages(by_vy):
    """Une ligne par rapport à partir de {millésime: {x: val}} : rapport récent
    en trait plein saillant, les autres en pointillé (dégradé `COLORS`). Aucun
    filtrage calendaire (l'axe x peut être une génération)."""
    out = []
    for vy in sorted(by_vy):
        s = by_vy[vy]
        pts = [{"x": x, "y": s[x]} for x in sorted(s)]
        if pts:
            out.append({"label": f"Rapport {vy}", "color": COLORS.get(vy, "#888888"),
                        "kind": "solid" if vy == LATEST else "dash", "points": pts})
    return out


_PROJ_KEYS = ("central", "référence", "reference", "tous scénarios", "sc. ref", "projection")


def _obs_proj(rows, scale=1.0):
    """Extrait (observé, projeté) d'une figure 'observé puis projeté'."""
    ym = _ymap(rows)

    def match(r, keys):
        for j in (1, 2):
            v = str(r[j]).lower().strip() if len(r) > j and r[j] is not None else ""
            if any(k in v for k in keys):
                return True
        return False

    obs = proj = None
    for r in rows:
        if obs is None and match(r, ("observé", "obs", "observations")):
            obs = {ym[i]: round(r[i] * scale, 3) for i in ym
                   if i < len(r) and isinstance(r[i], (int, float))}
        if proj is None and match(r, _PROJ_KEYS):
            proj = {ym[i]: round(r[i] * scale, 3) for i in ym
                    if i < len(r) and isinstance(r[i], (int, float))}
    return obs or {}, proj or {}


def _block_sub(rows, label_key, sub):
    """Sous-ligne col2==sub du 1er bloc dont col1 contient label_key."""
    ym = _ymap(rows)
    start = None
    for i, r in enumerate(rows):
        if len(r) > 1 and r[1] and label_key.lower() in str(r[1]).lower():
            start = i
            break
    if start is None:
        return {}
    for r in rows[start:start + 7]:
        if len(r) > 2 and str(r[2]).strip() == sub:
            return {ym[i]: r[i] for i in ym if i < len(r) and isinstance(r[i], (int, float))}
    return {}


def _ratio(a, b, scale=1.0):
    return {y: round(a[y] / b[y] * scale, 3) for y in a if y in b and b[y]}


def _series(obs, proj, color, obs_from=2000, to=2070):
    """Construit [observé solide, projeté pointillé] filtré sur [obs_from, to]."""
    o = [{"x": y, "y": obs[y]} for y in sorted(obs) if obs_from <= y <= to]
    p = [{"x": y, "y": proj[y]} for y in sorted(proj) if obs_from <= y <= to]
    out = []
    if o:
        out.append({"label": "Observé", "color": "#1f2d3d", "kind": "solid", "points": o})
    if p:
        out.append({"label": f"Projeté (réf. {LATEST})", "color": color, "kind": "dash", "points": p})
    return out


def _bounds(series, xpad=0, ypad=0.06):
    xs = [pt["x"] for s in series for pt in s["points"]]
    ys = [pt["y"] for s in series for pt in s["points"]]
    yr = (max(ys) - min(ys)) or 1
    return {"xMin": min(xs), "xMax": max(xs),
            "yMin": round(min(ys) - yr * ypad, 2), "yMax": round(max(ys) + yr * ypad, 2)}


def _popact(rows, own_year=None):
    """(obs, proj) du taux de croissance de la population active, en %.

    La figure superpose une ligne observée et une (ou plusieurs) ligne(s)
    projetée(s), chacune annotée du rapport dont elle provient (« (Rapport
    COR 2026) »). `own_year` : si fourni, on retient la projection de ce
    rapport précis ; sinon la première projection rencontrée."""
    ym = _ymap(rows)
    if not ym:
        return {}, {}

    def vals(r):
        return {ym[i]: round(r[i] * 100, 3) for i in ym
                if i < len(r) and isinstance(r[i], (int, float))}

    obs, proj, proj_rows = {}, {}, []
    for r in rows:
        lab = next((str(c) for c in r[1:3] if isinstance(c, str) and len(c) > 3), "")
        low = lab.lower()
        if "population active" not in low and "pop act" not in low:
            continue
        if "observ" in low and not obs:
            obs = vals(r)
        elif "projet" in low:
            proj_rows.append((low, vals(r)))
    if own_year:
        proj = next((v for lab, v in proj_rows
                     if f"rapport cor {own_year}" in lab or f"rapport {own_year}" in lab), {})
    if not proj and proj_rows:
        proj = proj_rows[0][1]
    return obs, proj


def extract_pop_active(path, own_year=None):
    """(obs, proj) du taux de croissance de la population active d'un rapport."""
    if not path:
        return {}, {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_title(wb, ["taux de croissance", "population active"])
    if ws is None:
        wb.close()
        return {}, {}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _popact(rows, own_year)


# Couleurs des grands groupes de régimes (Fig 2.6). On apparie par sous-chaîne
# du libellé pour rester robuste aux variations d'orthographe d'un rapport à
# l'autre (« Régimes spéciaux » avec ou sans espace final, etc.).
REGIME_COLORS = [
    ("complémentaires", "#9c27b0"),
    ("spéciaux", "#c2185b"),
    ("non-salariés", "#e8731c"),
    ("cnracl", "#6aa84f"),
    ("fpe", "#2f6fb0"),
    ("lura", "#1f4e79"),
]


def extract_dep_regimes(rows):
    """Une série par groupe de régimes (% du PIB), bloc « Hors transferts
    internes » de la figure des dépenses par groupe de régimes."""
    ym = _ymap(rows)
    if not ym:
        return []
    series, started = [], False
    for r in rows:
        c1 = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if not c1:
            continue
        if c1.lower().startswith("hors transferts"):
            started = True
            continue
        if started and c1.lower().startswith("y compris"):
            break
        if started:
            pts = [{"x": ym[i], "y": round(r[i] * 100, 3)} for i in ym
                   if i < len(r) and isinstance(r[i], (int, float))]
            if pts:
                col = next((c for k, c in REGIME_COLORS if k in c1.lower()), "#888888")
                series.append({"label": c1, "color": col, "kind": "solid", "points": pts})
    return series


# Composantes des ressources du système (Fig 2.11), appariées par sous-chaîne.
RESSOURCE_COLORS = [
    ("cotisations", "#1f4e79"),
    ("contribution d'équilibre", "#2f6fb0"),
    ("itaf", "#c2185b"),
    ("subventions", "#e8731c"),
    ("transferts", "#6aa84f"),
    ("autres", "#9c27b0"),
]

_COMPO_STOP = ("lecture", "champ", "source", "note")


def extract_composition(rows, scale=100, colors=None):
    """Une série par ligne libellée d'une figure de composition (chaque ligne
    une part qui, sommée aux autres, fait 100 %). On garde les lignes dont le
    libellé n'est pas une mention de bas de figure et qui portent ≥5 points."""
    ym = _ymap(rows)
    if not ym:
        return []
    colors = colors or []
    series = []
    for r in rows:
        c1 = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if not c1 or c1.lower().startswith(_COMPO_STOP):
            continue
        pts = [{"x": ym[i], "y": round(r[i] * scale, 3)} for i in ym
               if i < len(r) and isinstance(r[i], (int, float))]
        if len(pts) >= 5:
            col = next((c for k, c in colors if k in c1.lower()), "#888888")
            series.append({"label": c1, "color": col, "kind": "solid", "points": pts})
    return series


# --------------------------------------------------------------------------
# Figures à axe catégoriel (barres groupées / empilées) — modèle data-driven
# de l'explorateur : chaque indicateur porte `categories` (libellés de l'axe x)
# et des séries dont les points {x, y} ont x = index de catégorie.
# --------------------------------------------------------------------------
_CAT_STOP = ("lecture", "champ", "source", "note", "en ", "par ordre")


def _header_cols(rows, floor_col=2):
    """1re ligne portant ≥2 cellules texte à partir de `floor_col` : sert
    d'en-tête (catégories en colonnes). Renvoie [(index_colonne, libellé)]."""
    for r in rows[:8]:
        cells = [(i, str(c).strip()) for i, c in enumerate(r)
                 if i >= floor_col and isinstance(c, str) and c.strip()]
        if len(cells) >= 2:
            return cells
    return []


def extract_stacked(rows, scale=1.0, colors=None,
                    total_keys=("rdb", "revenu disponible", "total")):
    """Barres empilées signées : composants en lignes (col 1), catégories en
    colonnes (en-tête). La ligne « total » devient une série `total` (repère,
    hors empilement). Renvoie (categories, series)."""
    hdr = _header_cols(rows)
    if not hdr:
        return [], []
    idxs = [i for i, _ in hdr]
    cats = [t for _, t in hdr]
    colors = colors or []
    series, k = [], 0
    for r in rows:
        lab = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if not lab or lab.lower().startswith(_CAT_STOP):
            continue
        vals = [r[i] if i < len(r) and isinstance(r[i], (int, float)) else None for i in idxs]
        if all(v is None for v in vals):
            continue
        pts = [{"x": j, "y": round(v * scale, 3)} for j, v in enumerate(vals) if v is not None]
        is_total = any(tk in lab.lower() for tk in total_keys)
        col = colors[k % len(colors)] if colors and not is_total else "#1f2d3d"
        s = {"label": lab, "color": col, "points": pts}
        if is_total:
            s["total"] = True
        else:
            k += 1
        series.append(s)
    return cats, series


def extract_country_grouped(rows, scale=1.0, colors=None):
    """Barres groupées : catégories = lignes (col 1, ex. pays), séries =
    colonnes de l'en-tête (ex. tranches d'âge). Renvoie (categories, series)."""
    hdr = _header_cols(rows)
    if not hdr:
        return [], []
    idxs = [i for i, _ in hdr]
    labels = [t for _, t in hdr]
    colors = colors or ["#1f4e79", "#e8731c", "#6aa84f", "#c2185b", "#9c27b0"]
    cats, cols = [], [[] for _ in idxs]
    for r in rows:
        c = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if not c or c.lower().startswith(_CAT_STOP):
            continue
        vals = [r[i] if i < len(r) and isinstance(r[i], (int, float)) else None for i in idxs]
        if all(v is None for v in vals):
            continue
        ci = len(cats)
        cats.append(c)
        for j, v in enumerate(vals):
            if v is not None:
                cols[j].append({"x": ci, "y": round(v * scale, 3)})
    series = [{"label": labels[j], "color": colors[j % len(colors)], "points": cols[j]}
              for j in range(len(idxs)) if cols[j]]
    return cats, series


def extract_determinants(rows, horizon="2070", scale=1.0):
    """Cascade : décomposition de l'écart de dépenses (en points de PIB) entre
    deux rapports, à un horizon donné. Contributions signées + barre « résidu »
    pour boucler sur le total (les « Dont » ne somment pas exactement), puis la
    barre « total ». Renvoie (categories, series) ; le point total porte total=True."""
    hcol = None
    for r in rows[:6]:
        for i, c in enumerate(r):
            if str(c).strip() == horizon:
                hcol = i
                break
        if hcol is not None:
            break
    if hcol is None:
        return [], []

    def num(r):
        v = r[hcol] if hcol < len(r) else None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            t = v.strip().replace(",", ".")
            try:
                return float(t)
            except ValueError:
                return None
        return None

    SHORT = [("fécondité", "Fécondité"), ("solde migratoire", "Migrations"),
             ("espérance de vie", "Espérance de vie"), ("économique", "Hypothèses éco."),
             ("lfss", "Mesures LFSS 2026"), ("revalorisation", "Revalorisation")]
    total, contribs = None, []
    for r in rows:
        lab = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        ll = lab.lower()
        if ll.startswith("écart de dépenses"):
            total = num(r)
        elif ll.startswith("dont") and "démographi" not in ll:
            v = num(r)
            if v is None:
                continue
            short = next((s for k, s in SHORT if k in ll), lab.replace("Dont", "").strip())
            contribs.append((short, v))
    if total is None or not contribs:
        return [], []
    resid = round(total - sum(v for _, v in contribs), 3)
    cats = [c for c, _ in contribs] + ["Autres / résidu", "Écart total"]
    pts = [{"x": i, "y": round(v * scale, 3)} for i, (_, v) in enumerate(contribs)]
    pts.append({"x": len(contribs), "y": round(resid * scale, 3)})
    pts.append({"x": len(contribs) + 1, "y": round(total * scale, 3), "total": True})
    return cats, [{"label": "Écart de dépenses (points de PIB)", "color": "#1f4e79", "points": pts}]


# Profil par âge (Fig 1.8) : courbe catégorielle. Catégories = tranches d'âge,
# une série par (sexe × année). Couleur par année, plein = femmes, tireté = hommes.
PROFILE_YEAR_COLORS = {1975: "#9aa7b4", 2024: "#2f6fb0", 2025: "#2f6fb0",
                       2050: "#e8731c", 2070: "#c2185b"}


def _age_key(s):
    """Normalise un libellé de tranche d'âge (« 15-19 ans », « 15-19 » → « 15-19 »)."""
    return str(s).replace("ans", "").replace("\xa0", " ").strip()


def extract_age_profile(rows, scale=1.0):
    ycols = {}
    for r in rows[:6]:
        ic = [(i, c) for i, c in enumerate(r) if isinstance(c, int) and 1900 <= c <= 2100]
        if len(ic) >= 2:
            ycols = {i: c for i, c in ic}
            break
    if not ycols:
        return [], []
    cats, catidx, data, gender = [], {}, {}, None
    for r in rows:
        g = str(r[0]).strip() if len(r) > 0 and isinstance(r[0], str) else ""
        if g in ("Femmes", "Hommes"):
            gender = g
        b = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if not gender or not b or b.lower().startswith(("année", "champ", "source", "note", "lecture")):
            continue
        vals = {yr: r[i] for i, yr in ycols.items() if i < len(r) and isinstance(r[i], (int, float))}
        if not vals:
            continue
        key = _age_key(b)
        if key not in catidx:
            catidx[key] = len(cats)
            cats.append(key)
        ci = catidx[key]
        for yr, v in vals.items():
            data.setdefault((gender, yr), {})[ci] = round(v * scale, 3)
    series = []
    for yr in sorted({yr for (_, yr) in data}):
        for gender, kind in (("Femmes", "solid"), ("Hommes", "dash")):
            d = data.get((gender, yr))
            if not d:
                continue
            series.append({"label": f"{gender} {yr}", "color": PROFILE_YEAR_COLORS.get(yr, "#888888"),
                           "kind": kind, "points": [{"x": i, "y": d[i]} for i in sorted(d)]})
    return cats, series


# Taux d'emploi des seniors (Fig 4.1) : trois tranches quinquennales, chacune
# ventilée Ensemble/Femmes/Hommes. Une teinte par tranche, intensité par sexe.
AGE_BAND_COLORS = {
    "55-59": {"Ensemble": "#1f4e79", "Femmes": "#3d7bbf", "Hommes": "#8fb8e0"},
    "60-64": {"Ensemble": "#b34700", "Femmes": "#e8731c", "Hommes": "#f4a96b"},
    "65-69": {"Ensemble": "#2e6b1f", "Femmes": "#6aa84f", "Hommes": "#a6c98f"},
}


def extract_age_bands(rows, scale=100):
    """Une série continue par couple (tranche d'âge × sexe) d'une figure en blocs.

    Chaque tranche apparaît en deux blocs (observé puis projeté), chacun introduit
    par une ligne d'en-tête de tranche (libellé contenant « ans », ex. « 55-59 ans »)
    et sa propre ligne d'années, suivies des lignes Ensemble / Femmes / Hommes. On
    fusionne les deux blocs en une seule courbe par couple (l'observé l'emporte sur
    les années en recouvrement). La ligne de titre (qui contient aussi « ans ») est
    sans effet : elle est écrasée avant la première ligne de données."""
    if rows is None:
        return []
    acc, order, band, ym = {}, [], None, None
    for r in rows:
        bh = next((c.strip() for c in r if isinstance(c, str)
                   and "ans" in c.lower() and any(ch.isdigit() for ch in c)), None)
        if bh:
            band, ym = bh.split()[0], None  # « 55-59 ans » → « 55-59 »
            continue
        yc = [(i, c) for i, c in enumerate(r) if isinstance(c, int) and 1950 <= c <= 2100]
        if len(yc) >= 3:
            ym = {i: c for i, c in yc}
            continue
        if not (band and ym):
            continue
        lab = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if lab in ("Ensemble", "Femmes", "Hommes"):
            key = (band, lab)
            d = acc.get(key)
            if d is None:
                d = {}
                acc[key], _ = d, order.append(key)
            for i, y in ym.items():
                if i < len(r) and isinstance(r[i], (int, float)) and y not in d:
                    d[y] = round(r[i] * scale, 3)
    series = []
    for band, lab in order:
        d = acc[(band, lab)]
        if d:
            col = AGE_BAND_COLORS.get(band, {}).get(lab, "#888888")
            series.append({"label": f"{band} ans — {lab}", "color": col, "kind": "solid",
                           "points": [{"x": x, "y": d[x]} for x in sorted(d)]})
    return series


def gendered_cohort(path, sheet_keys, pairs, scale=1.0, lo=None, hi=None,
                    floor=1900, exclude=()):
    """Deux courbes (Femmes/Hommes) d'une figure à en-tête d'années horizontal.

    `pairs` = [(label, couleur, [clés-de-ligne…]), …] ; les clés d'un même couple
    sont fusionnées (observé + projeté). Filtre de plausibilité optionnel."""
    out = []
    for label, color, keys in pairs:
        s = {}
        for k in keys:
            s.update(pick_cohort_row(path, sheet_keys, (k,), scale, floor, exclude=exclude))
        if lo is not None:
            s = _plausible(s, lo, hi)
        if s:
            out.append({"label": label, "color": color, "kind": "solid",
                        "points": [{"x": x, "y": s[x]} for x in sorted(s)]})
    return out


# Palette générique pour les séries multiples (régimes, pays, composants…).
DEFAULT_PALETTE = ["#1f4e79", "#e8731c", "#6aa84f", "#c2185b", "#9c27b0",
                   "#2f6fb0", "#b34700", "#0f7b6c", "#7d8ca0", "#8a6d3b"]

# Couleurs par nombre d'enfants (figures « droits familiaux » par génération).
ENF_COLORS = [("sans enf", "#1f4e79"), ("1 ou 2 enf", "#e8731c"),
              ("3 enf", "#c2185b"), ("ensemble", "#7d8ca0")]

_TS_STOP = ("lecture", "champ", "source", "note", "ensemble des", "par ordre", "année")


def _year_of(c, floor):
    if isinstance(c, int) and floor <= c <= 2100:
        return c
    if isinstance(c, str) and c.strip().isdigit() and floor <= int(c) <= 2100:
        return int(c)
    return None


def _num(c):
    if isinstance(c, (int, float)):
        return float(c)
    if isinstance(c, str):
        try:
            return float(c.strip().replace("\xa0", "").replace(",", "."))
        except ValueError:
            return None
    return None


def extract_rows_timeseries(rows, scale=1.0, floor=1990, colors=None, label_col=1, stop=_TS_STOP):
    """Multi-courbes horizontales : une série par ligne libellée (col `label_col`),
    années en en-tête (entières ou texte). Généralise extract_dep_regimes sans bornage."""
    ycols, hdr_idx = {}, -1
    for idx, r in enumerate(rows[:8]):
        ic = [(i, y) for i, c in enumerate(r) if (y := _year_of(c, floor)) is not None]
        if len(ic) >= 3:
            ycols, hdr_idx = dict(ic), idx
            break
    if not ycols:
        return []
    colors = colors or DEFAULT_PALETTE
    series, k = [], 0
    for idx, r in enumerate(rows):
        if idx == hdr_idx:
            continue
        lab = str(r[label_col]).strip() if len(r) > label_col and isinstance(r[label_col], str) else ""
        if not lab or lab.lower().startswith(stop):
            continue
        pts = [{"x": ycols[i], "y": round(v * scale, 3)} for i in ycols
               if i < len(r) and (v := _num(r[i])) is not None]
        if len(pts) < 3:
            continue
        series.append({"label": lab, "color": colors[k % len(colors)], "kind": "solid", "points": pts})
        k += 1
    return series


def extract_cols_timeseries(rows, spec, scale=1.0, floor=1960, year_col=1):
    """Multi-courbes verticales : années descendant la colonne `year_col`, séries
    décrites explicitement par `spec` = [(index_colonne, libellé, couleur), …]
    (robuste aux en-têtes sur deux niveaux)."""
    out = []
    for ci, label, color in spec:
        pts = []
        for r in rows:
            y = _year_of(r[year_col], floor) if len(r) > year_col else None
            if y is None:
                continue
            v = _num(r[ci]) if ci < len(r) else None
            if v is not None:
                pts.append({"x": y, "y": round(v * scale, 3)})
        if len(pts) >= 3:
            out.append({"label": label, "color": color, "kind": "solid", "points": pts})
    return out


def extract_cohort_groups(rows, scale=1.0, floor=1900):
    """Courbes par génération à deux niveaux : sexe (rémanent, repéré dans les
    colonnes à gauche du libellé), sous-catégorie dans la colonne juste avant les
    générations (ex. nombre d'enfants). Trait plein = femmes, tireté = hommes ;
    couleur par sous-catégorie. Tolère le décalage de colonnes d'une figure à l'autre."""
    ycols = {}
    for r in rows[:6]:
        ic = [(i, y) for i, c in enumerate(r) if (y := _year_of(c, floor)) is not None]
        if len(ic) >= 3:
            ycols = dict(ic)
            break
    if not ycols:
        return []
    sub_col = min(ycols) - 1
    series, gender = [], None
    for r in rows:
        for j in range(sub_col):
            c = r[j] if j < len(r) else None
            if isinstance(c, str) and c.strip().lower().startswith(("femme", "homme")):
                gender = c.strip()
        sub = str(r[sub_col]).strip() if len(r) > sub_col and isinstance(r[sub_col], str) else ""
        if not gender or not sub or sub.lower().startswith(("année", "champ", "source", "note", "lecture")):
            continue
        pts = [{"x": ycols[i], "y": round(v * scale, 3)} for i in ycols
               if i < len(r) and (v := _num(r[i])) is not None]
        if len(pts) < 3:
            continue
        col = next((c for k, c in ENF_COLORS if k in sub.lower()), "#888888")
        kind = "solid" if gender.lower().startswith("femme") else "dash"
        series.append({"label": f"{gender} — {sub}", "color": col, "kind": kind, "points": pts})
    return series


def _short_regime(c):
    """Raccourcit les libellés de régimes (Fig droits familiaux) pour l'axe x."""
    cl = c.lower()
    for k, s in [("base", "Base (privé/agricole/indép.)"), ("complémentaire", "Complémentaires"),
                 ("fonction publique", "Fonction publique"), ("spéci", "Régimes spéciaux"),
                 ("non-salari", "Non-salariés"), ("ensemble", "Ensemble")]:
        if k in cl:
            return s
    return " ".join(c.split())


# Grands régimes retenus (les ~30 régimes du thématique 2017 seraient illisibles).
REG_KEEP = ["CNAV", "AGIRC-ARRCO", "FPE", "CNRACL", "IRCANTEC", "RSI",
            "MSA salariés agricoles"]
_REG_NOISE = ("écart", "partage", "taux de chômage", "population", "dont")


def extract_regime_series(rows, sub_key, scale=1.0, floor=2000, colors=None, keep=REG_KEEP):
    """Une courbe par grand régime : régime rémanent en col 1, sous-ligne
    sélectionnée par `sub_key` (col 2, ex. « ensemble » ou scénario « 1,3 »),
    années en en-tête. Filtre `keep` (grands régimes) et déduplique (1re sous-ligne
    par régime, hors lignes d'écarts/variantes). Thématique 2017 « par régime »."""
    ycols, hdr_idx = {}, -1
    for idx, r in enumerate(rows[:6]):
        ic = [(i, y) for i, c in enumerate(r) if (y := _year_of(c, floor)) is not None]
        if len(ic) >= 3:
            ycols, hdr_idx = dict(ic), idx
            break
    if not ycols:
        return []
    colors = colors or DEFAULT_PALETTE
    series, regime, seen, k = [], None, set(), 0
    for idx, r in enumerate(rows):
        if idx == hdr_idx:
            continue
        reg = str(r[1]).strip() if len(r) > 1 and isinstance(r[1], str) else ""
        if reg:
            regime = reg
        sub = str(r[2]).strip().lower() if len(r) > 2 and isinstance(r[2], str) else ""
        if not regime or sub_key not in sub or sub.startswith(_REG_NOISE):
            continue
        if keep and regime not in keep:
            continue
        if regime in seen:
            continue
        pts = [{"x": ycols[i], "y": round(v * scale, 3)} for i in ycols
               if i < len(r) and (v := _num(r[i])) is not None]
        if len(pts) < 3:
            continue
        seen.add(regime)
        series.append({"label": regime, "color": colors[k % len(colors)], "kind": "solid", "points": pts})
        k += 1
    return series


def extract_category_snapshot(rows, row_key="par pays", scale=1.0, accent="france",
                              accent_color="#c2185b", base_color="#5b6f93"):
    """Instantané par catégorie : catégories = en-tête (ex. pays en colonnes), une
    barre par catégorie issue de la ligne `row_key`. Tri décroissant, catégorie
    `accent` (France) en couleur saillante (couleur portée par chaque point)."""
    hdr = _header_cols(rows)
    if not hdr:
        return [], []
    datarow = next((r for r in rows if len(r) > 1 and isinstance(r[1], str)
                    and row_key in r[1].strip().lower()), None)
    if datarow is None:
        return [], []
    pairs = []
    for i, name in hdr:
        v = _num(datarow[i]) if i < len(datarow) else None
        if v is not None:
            pairs.append((" ".join(str(name).split()), round(v * scale, 3)))
    if not pairs:
        return [], []
    pairs.sort(key=lambda kv: kv[1], reverse=True)
    cats = [p[0] for p in pairs]
    pts = [{"x": j, "y": v, "color": accent_color if accent in cats[j].lower() else base_color}
           for j, (_, v) in enumerate(pairs)]
    return cats, [{"label": "Par pays", "color": base_color, "points": pts}]


def build_explorer(multi=None):
    """`multi` : séries multi-millésimes calculées dans build() —
    {indicateur: ({vy: {année: val}} projections, obs {année: val})}."""
    multi = multi or {}
    ind = {}

    def add(iid, label, unit, suffix, series, desc, source, obs_from=2000,
            xLabel="Année", chartType="line", barMode=None, categories=None,
            waterfall=False, yMin=None, yMax=None):
        series = [s for s in series if s["points"]]
        if not series:
            print("✗ explorateur:", iid, "(vide)")
            return None
        rec = {"label": label, "unit": unit, "suffix": suffix, "xLabel": xLabel,
               "desc": desc, "source": source, "series": series}
        if categories is not None:
            rec["categories"] = categories
        if chartType == "bar":
            # Axe catégoriel : pas de bornes x ; les bornes y sont laissées au
            # moteur (qui inclut les cumuls empilés) sauf si on les force.
            rec["chartType"] = "bar"
            rec["barMode"] = barMode or "grouped"
            if waterfall:
                rec["waterfall"] = True
            if yMin is not None:
                rec["yMin"] = yMin
            if yMax is not None:
                rec["yMax"] = yMax
        else:
            # Courbe : bornes x/y. En catégoriel, x = index 0..n−1.
            rec.update(_bounds(series))
        ind[iid] = rec
        return iid

    def vintage_series(obs, projs, obs_from=2000, notes=None):
        """Observé (trait plein) + une projection en pointillé par millésime.

        `notes` : annotation de légende par millésime (ex. l'hypothèse retenue,
        « Rapport 2023 (4,5 %) »)."""
        out = []
        if obs:
            out.append({"label": "Observé", "color": "#1f2d3d", "kind": "solid",
                        "points": [{"x": y, "y": obs[y]} for y in sorted(obs)
                                   if obs_from <= y <= 2070]})
        for vy in sorted(projs):
            p = projs[vy]
            pts = [{"x": y, "y": p[y]} for y in sorted(p) if int(vy) <= y <= 2070]
            if pts:
                note = (notes or {}).get(vy)
                label = f"Rapport {vy} ({note})" if note else f"Rapport {vy}"
                out.append({"label": label, "color": COLORS.get(vy, "#888888"),
                            "kind": "dash", "points": pts})
        return out

    NOTES = multi.get("_notes", {})

    DEMO, PENS, FIN, ECO = "#2f6fb0", "#c2185b", "#e8731c", "#6aa84f"

    # --- Démographie
    r = _rows("partie 1", "Fig 1.2")
    if r:
        o, p = _obs_proj(r, 0.001)
        series = _series(o, p, DEMO, 1995)
        if "migration" in multi:
            series = vintage_series(o, multi["migration"], 1995)
        add("migration", "Solde migratoire", "milliers de personnes / an", " k",
            series,
            "Le solde migratoire (entrées − sorties). Hypothèse de +70 000/an jusqu'au "
            "rapport 2025, relevée à +150 000/an dans les projections 2026.",
            "COR / INSEE, rapports 2017-2026 (solde migratoire, scénario central de chaque rapport).", 1995)
    r = _rows("partie 1", "Fig 1.11")
    if r:
        o, p = _obs_proj(r, 100)
        series = _series(o, p, DEMO)
        if multi.get("chomage"):
            series = vintage_series(o, multi["chomage"], 2000, NOTES.get("chomage"))
        add("chomage", "Taux de chômage", "%", " %", series,
            "Le taux de chômage de long terme retenu par chaque rapport — longtemps 7 %, "
            "abaissé à 4,5 % en 2023 (cible de plein emploi) puis 5 % en 2024, avant de "
            "revenir à 7 % dès 2025.",
            "COR, rapports 2017-2026 (taux de chômage observé puis projeté).")
    r = _rows("partie 1", "Fig 1.5")
    if r:
        o, p = _obs_proj(r, 1)
        series = _series(o, p, DEMO)
        if multi.get("ratio_demo"):
            series = vintage_series(o, multi["ratio_demo"])
        add("ratio_demo", "Rapport démographique (20-64 ans / 65 ans et +)", "ratio", "",
            series,
            "Combien de personnes en âge de travailler pour une personne de 65 ans et plus. "
            "Il s'effondre avec le vieillissement — un peu plus à chaque jeu de projections.",
            "COR / INSEE, rapports 2019-2026 (rapport démographique, projections INSEE successives).")
    # Les effectifs cotisants/retraités ne sont publiés que dans les « données
    # complémentaires », absentes du rapport 2026 : on superpose 2024 et 2025.
    cr = _rows("complémentaires", "Cotisants_Retraités", R25)
    if cr:
        co = _block_sub(cr, "cotisants", "Obs")
        ro = _block_sub(cr, "retraités", "Obs")
        obs_ratio = _ratio(co, ro)
        if multi.get("cot_ret"):
            series = vintage_series(obs_ratio, multi["cot_ret"])
        else:
            cp = _block_sub(cr, "cotisants", "Sc. Ref")
            rp = _block_sub(cr, "retraités", "Sc. Ref")
            series = _series(obs_ratio, _ratio(cp, rp), DEMO)
        add("cot_ret", "Nombre de cotisants par retraité", "ratio", "",
            series,
            "Le cœur du système par répartition : chaque retraité est financé par les "
            "cotisations des actifs. Ce ratio baisse de ~1,8 vers ~1,4.",
            "COR, rapports 2024-2025 (données complémentaires, non republiées en 2026).")
    fec = _rows("partie 1", "Fig 1.1")
    if fec:
        o = _row_label(fec, "Observé")
        o.update(_row_label(fec, "Données provisoires"))
        p = _row_label(fec, "central")
        series = _series(o, p, DEMO)
        if "fecondite" in multi:
            series = vintage_series(o, multi["fecondite"])
        add("fecondite", "Indice de fécondité", "enfants / femme", "",
            series,
            "Nombre d'enfants par femme. Hypothèse de 1,95 jusqu'en 2021, abaissée à "
            "1,80 en 2022 puis à 1,45 en 2026 — cette fois sous l'observé (~1,56 en 2025).",
            "COR / INSEE, rapports 2017-2026 (fécondité, scénario central de chaque rapport).")
    ev = _rows("partie 1", "Fig 1.3")
    if ev:
        o = _row_label(ev, "Observé")
        p = _row_label(ev, "scénario central")
        series = _series(o, p, DEMO)
        if "esp_vie" in multi:
            series = vintage_series(o, multi["esp_vie"])
        add("esp_vie", "Espérance de vie à 65 ans (femmes)", "ans", " ans",
            series,
            "Nombre d'années encore à vivre à 65 ans (femmes). Elle progresse à chaque "
            "jeu de projections, ce qui allonge la durée de retraite.",
            "COR / INSEE, rapports 2019-2026 (espérance de vie à 65 ans, scénario central de chaque rapport).")

    # --- Emploi & économie
    r = _rows("partie 1", "Fig 1.12")
    if r:
        o, p = _obs_proj(r, 100)
        series = _series(o, p, ECO)
        if multi.get("emploi"):
            series = vintage_series(o, multi["emploi"], 2000, NOTES.get("emploi"))
        add("emploi", "Taux d'emploi des 15-64 ans", "%", " %", series,
            "Part des 15-64 ans qui ont un emploi. Plus il est élevé, plus il y a de "
            "cotisants.",
            "COR, rapports 2024-2026 (taux d'emploi observé puis projeté).")
    r = _rows("partie 1", "Fig 1.6")
    if r:
        obs, _p = _popact(r)
        series = _series(obs, _p, ECO)
        if multi.get("pop_active"):
            series = vintage_series(obs, multi["pop_active"])
        add("pop_active", "Croissance de la population active", "%/an", " %", series,
            "Le rythme d'évolution de la population active (personnes en emploi ou qui en "
            "cherchent) : c'est la base des cotisants qui s'élargit ou se resserre. Chaque "
            "rapport projette un net ralentissement, révisé d'une édition à l'autre.",
            "COR / INSEE, rapports 2023-2026 (taux de croissance de la population active, "
            "scénario de référence de chaque rapport).")
    pr = _rows("partie 1", "Fig 1.10")
    if pr:
        annual = {}
        for r2 in pr:
            if len(r2) > 1 and str(r2[1]).lower().startswith("croissance annuelle observ"):
                ym = _ymap(pr)
                annual = {ym[i]: r2[i] * 100 for i in ym if i < len(r2) and isinstance(r2[i], (int, float))}
                break
        ma = {p["x"]: p["y"] for p in moving_average(annual, 5)}
        if multi.get("prod_path"):
            s = vintage_series(ma, multi["prod_path"], 2000, NOTES.get("prod_path"))
        else:
            s = _series(ma, _row_label(pr, "Scénario de référence", 100), ECO)
        if s:
            s[0]["label"] = "Observé (moy. mobile 5 ans)"
            iid = add("productivite", "Productivité du travail", "%/an", " %", s,
                      "Croissance de la productivité : moteur des salaires donc des cotisations. "
                      "Le scénario central est passé de 1,3 % à 1,0 % (2023) puis 0,7 % (2025). "
                      "Le scénario du rapport 2020 s'ouvrait sur l'effondrement (−8,7 % en 2020) "
                      "puis le rebond (+8,4 % en 2021) du Covid : ces valeurs extrêmes sont "
                      "affichées au-delà des coupures de l'axe (échelle interrompue).",
                      "COR, rapports 2018-2026 (productivité observée puis projetée, scénario central de chaque rapport).")
            if iid:
                # Axe borné pour garder lisibles les hypothèses de long terme ;
                # la pointe Covid du rapport 2020, très hors échelle, est
                # affichée dans les bandes à axe interrompu (voir chart.js).
                ind[iid]["yMin"], ind[iid]["yMax"] = -1.0, 2.5

    # --- Pensions & retraités
    r = _rows("synthèse", "Âge conjoncturel")
    if r:
        o, p = _obs_proj(r, 1)
        series = _series(o, p, PENS)
        if "age" in multi:
            series = vintage_series(o, multi["age"])
        add("age_depart", "Âge de départ à la retraite", "ans", " ans", series,
            "L'âge « conjoncturel » de départ : il monte sous l'effet des réformes. "
            "Chaque rapport reflète la législation du moment (réforme de 2023, "
            "aménagements de la LFSS 2026).",
            "COR / DREES, rapports 2023-2026 (âge conjoncturel).")
    p24 = _rows("partie 2", "Fig 2.4")
    if p24:
        pen = _ratio(_block_sub(p24, "Pension moyenne", "Obs"), _block_sub(p24, "Rémunération nett", "Obs"), 100)
        penp = _ratio(_block_sub(p24, "Pension moyenne", "Sc. Ref"), _block_sub(p24, "Rémunération nett", "Sc. Ref"), 100)
        desc = "La pension nette moyenne en % du salaire net moyen. Elle décroche au fil de la projection."
        if penp:
            y0, y1 = min(penp), max(penp)
            desc = (f"La pension nette moyenne en % du salaire net moyen. Elle décroche : "
                    f"~{penp[y0]:.0f} % aujourd'hui, ~{penp[y1]:.0f} % en {y1}.")
        series = _series(pen, penp, PENS)
        if multi.get("pension_rel"):
            series = vintage_series(pen, multi["pension_rel"])
        add("pension_rel", "Pension moyenne rapportée au salaire net", "%", " %",
            series, desc,
            "COR, rapports 2022-2026 (pension nette moyenne et revenu net moyen, "
            "scénario de référence de chaque rapport).")
    nv = _rows("synthèse", "Niveau de vie relatif")
    if nv:
        o, p = _obs_proj(nv, 100)
        series = _series(o, p, PENS, 1996)
        if "nv" in multi:
            series = vintage_series(o, multi["nv"], 1996)
        add("niveau_vie", "Niveau de vie des retraités / population", "%", " %",
            series,
            "Niveau de vie moyen des retraités rapporté à l'ensemble de la population "
            "(100 % = parité). Chaque rapport repousse un peu le décrochage projeté. "
            "Attention : l'Insee a révisé la série observée entre les rapports 2024 et "
            "2025 (≈2 points) — comparer les pentes plutôt que les niveaux.",
            "COR / INSEE-DGI, rapports 2023-2026.", 1996)

    # --- Finances
    for iid, sheet, label, desc, mkey in [
        ("depenses", "Dépenses en %", "Dépenses de retraite (% du PIB)",
         "Ce que le système verse, en part de la richesse nationale.", "depenses"),
        ("ressources", "Ressources en %", "Ressources du système (% du PIB)",
         "Ce que le système encaisse (cotisations, impôts affectés…).", "ressources"),
        ("solde", "Solde en %", "Solde du système (% du PIB)",
         "Ressources − dépenses. Négatif = déficit.", "solde"),
    ]:
        r = _rows("synthèse", sheet)
        if r:
            o, p = _obs_proj(r, 100)
            series = _series(o, p, FIN)
            src = "COR, rapport 2026 (synthèse, scénario de référence)."
            if mkey in multi:
                series = vintage_series(o, multi[mkey])
                src = ("COR, rapports 2016-2026 — scénario de référence de chaque rapport "
                       "(1,3 % avant 2023 ; conventions comptables : « COR » jusqu'en 2019, "
                       "EPR ensuite).")
            add(iid, label, "% du PIB", " %", series, desc, src)

    r = _rows("partie 2", "Fig 2.6")
    if r:
        series = extract_dep_regimes(r)
        add("dep_regimes", "Dépenses de retraite par groupe de régimes", "% du PIB", " %",
            series,
            "Comment se répartissent les dépenses de retraite (en % du PIB) entre les grands "
            "groupes de régimes : régime général et alignés (LURA), fonction publique (FPE, "
            "CNRACL), non-salariés, régimes spéciaux et régimes complémentaires. Vue du "
            "scénario de référence du rapport le plus récent.",
            "COR, rapport 2026 (dépenses par groupe de régimes, hors transferts internes, "
            "scénario de référence).")

    r = _rows("partie 2", "Fig 2.1")
    if r:
        obs = _row_label(r, "dépense publique", 100)
        series = [{"label": "Observé", "color": "#1f2d3d", "kind": "solid",
                   "points": [{"x": y, "y": obs[y]} for y in sorted(obs)]}] if obs else []
        add("dep_pub", "Part des retraites dans la dépense publique", "%", " %", series,
            "La part des dépenses de retraite dans l'ensemble de la dépense publique : "
            "près d'un quart. C'est ce que pèse le système rapporté à tout ce que dépense "
            "la puissance publique (et non au seul PIB).",
            "COR, rapports à la CCSS 2002-2025 ; comptabilité nationale Insee base 2020.")

    r = _rows("partie 2", "Fig 2.11")
    if r:
        series = extract_composition(r, 100, RESSOURCE_COLORS)
        add("struct_ressources", "Structure des ressources du système", "%", " %", series,
            "D'où vient l'argent du système de retraite : cotisations sociales (les deux "
            "tiers), impôts et taxes affectés (ITAF, en hausse), contributions et "
            "subventions d'équilibre de l'État, transferts. Les parts somment à 100 %.",
            "COR, rapport 2026 (structure des ressources de 2004 à 2025).")

    # --- Parties 3 & 4 : séries par génération / par année, superposées par rapport
    if multi.get("taux_rempl"):
        add("taux_rempl", "Taux de remplacement net à la liquidation", "%", " %",
            overlay_vintages(multi["taux_rempl"]),
            "Pension à la liquidation rapportée au dernier salaire, cas-type du salarié "
            "non-cadre du privé partant au taux plein, scénario de référence. Il baisse "
            "de génération en génération — et chaque rapport en redessine la pente.",
            "COR, rapports 2016-2026 (taux de remplacement net, cas-type non-cadre du "
            "privé, scénario de référence de chaque rapport).", xLabel="Génération")
    if multi.get("duree_retraite"):
        add("duree_retraite", "Durée de retraite par génération", "ans", " ans",
            overlay_vintages(multi["duree_retraite"]),
            "Nombre d'années passées à la retraite, en moyenne par génération (scénario "
            "central de mortalité). Elle s'allonge avec l'espérance de vie ; les révisions "
            "démographiques d'un rapport à l'autre la déplacent.",
            "COR, rapports 2016-2026 (durée de retraite, moyenne par génération, scénario "
            "central de mortalité).", xLabel="Génération")
    if multi.get("duree_carriere"):
        add("duree_carriere", "Durée de carrière par génération", "ans", " ans",
            overlay_vintages(multi["duree_carriere"]),
            "Nombre d'années de carrière validées, en moyenne par génération. C'est l'autre "
            "bout de l'équation : plus la carrière est longue, plus la retraite est tardive "
            "ou élevée.",
            "COR, rapports 2016-2026 (durée de carrière, moyenne par génération).",
            xLabel="Génération")
    if multi.get("age_moyen_depart"):
        add("age_moyen_depart", "Âge moyen de départ par génération", "ans", " ans",
            overlay_vintages(multi["age_moyen_depart"]),
            "Âge moyen effectif de départ à la retraite, par génération (observé puis "
            "projeté). Il remonte sous l'effet des réformes successives.",
            "COR, rapports 2023-2026 (âge moyen de départ par génération).",
            xLabel="Génération")
    r = _rows("partie 4", "Fig 4.1")
    if r:
        add("emploi_seniors", "Taux d'emploi des seniors par tranche d'âge", "%", " %",
            extract_age_bands(r, 100),
            "Part des seniors en emploi, par tranche quinquennale (55-59, 60-64, 65-69) et "
            "par sexe, depuis 1975. Un déterminant majeur de l'équilibre du système : plus "
            "de seniors en emploi, c'est plus de cotisants et moins de retraités précoces. "
            "Le taux progresse fortement depuis les années 2000, surtout chez les 55-59 ans.",
            "COR, rapport 2026 (taux d'emploi des 55-64 ans par tranche d'âge quinquennal).")
    if multi.get("pauvrete"):
        add("pauvrete", "Taux de pauvreté des retraités", "%", " %",
            overlay_vintages(multi["pauvrete"]),
            "Part des retraités sous le seuil de pauvreté (60 % du niveau de vie médian). "
            "Il reste nettement inférieur à celui de l'ensemble de la population, mais "
            "chaque rapport en prolonge et révise la mesure.",
            "COR / INSEE-ERFS, rapports 2016-2026 (taux de pauvreté de l'ensemble des "
            "retraités).")

    r = _rows("partie 3", "Fig 3.24")
    if r:
        rg = vertical_ratio(first_file(R26, "partie 3"), ("rapporté", "hommes"))
        rg = _plausible(rg, 40, 90)
        series = [{"label": "Femmes / hommes", "color": "#c2185b", "kind": "solid",
                   "points": [{"x": y, "y": rg[y]} for y in sorted(rg)]}] if rg else []
        add("ecart_genre", "Pension des femmes rapportée à celle des hommes", "%", " %",
            series,
            "Pension totale moyenne (y compris majorations) des femmes en % de celle des "
            "hommes, observée puis projetée. L'écart se réduit lentement — de ~70 % à un peu "
            "plus de 85 % attendu — sans jamais se refermer.",
            "COR, rapport 2026 (montant brut moyen de pension totale y compris majorations, "
            "femmes rapporté aux hommes, observé puis projeté).")

    # --- Parties 3 & 4 (suite) : nouveaux indicateurs récurrents
    if multi.get("tri"):
        add("tri", "Taux de rendement interne du système", "%", " %",
            overlay_vintages(multi["tri"]),
            "Ce que « rapporte » une carrière de cotisations : le taux qui égalise cotisations "
            "versées et pensions perçues, cas-type du salarié non-cadre du privé. Un indicateur "
            "officiel de suivi du COR, qui décline de génération en génération.",
            "COR, rapports 2016-2026 (taux de rendement interne net, cas-type non-cadre du "
            "privé, scénario de référence).", xLabel="Génération")
    if multi.get("taux_cotisation"):
        add("taux_cotisation", "Taux de cotisation retraite (cas-type)", "%", " %",
            overlay_vintages(multi["taux_cotisation"]),
            "Taux de cotisation retraite (tous régimes) appliqué au cas-type du salarié "
            "non-cadre du privé, par génération. Il mesure l'effort contributif qui finance "
            "les pensions.",
            "COR, rapports 2016-2026 (taux de cotisation retraite, tous régimes, cas-type "
            "non-cadre du privé).", xLabel="Génération")
    if multi.get("intensite_pauvrete"):
        add("intensite_pauvrete", "Intensité de la pauvreté des retraités", "%", " %",
            overlay_vintages(multi["intensite_pauvrete"]),
            "Quand un retraité est pauvre, à quel point l'est-il : écart relatif entre son "
            "niveau de vie et le seuil de pauvreté. Complète le taux de pauvreté, qui n'en "
            "compte que la fréquence.",
            "COR / INSEE-ERFS, rapports 2016-2026 (intensité de la pauvreté, ensemble des "
            "retraités).")

    p3 = first_file(R26, "partie 3")
    series = gendered_cohort(p3, ("durée moyenne d'assurance", "femmes"),
                             [("Femmes", "#c2185b", ["femmes"]),
                              ("Hommes", "#1f4e79", ["hommes"])],
                             scale=1, lo=100, hi=200)
    add("assurance_genre", "Durée d'assurance validée femmes / hommes", "trimestres",
        " trim.", series,
        "Nombre de trimestres validés en moyenne, par sexe et par génération. L'écart se "
        "comble peu à peu — carrières féminines plus longues — mais reste sensible.",
        "COR, rapport 2026 (durée moyenne d'assurance des femmes et des hommes, par "
        "génération).", xLabel="Génération")

    series = gendered_cohort(p3, ("âge moyen de départ", "femmes"),
                             [("Femmes", "#c2185b", ["femmes (observé)", "femmes (projeté)"]),
                              ("Hommes", "#1f4e79", ["hommes (observé)", "hommes (projeté)"])],
                             scale=1, lo=55, hi=70)
    add("age_depart_genre", "Âge moyen de départ femmes / hommes", "ans", " ans", series,
        "Âge moyen effectif de départ à la retraite, par sexe et par génération (observé puis "
        "projeté). Les femmes partent en moyenne un peu plus tard, l'écart se resserrant avec "
        "les réformes.",
        "COR, rapport 2026 (âge moyen de départ des femmes et des hommes, par génération).",
        xLabel="Génération")

    r = _rows("partie 2", "Fig 2.15")
    if r:
        add("solde_regimes", "Solde projeté par groupe de régimes", "% du PIB", " %",
            extract_dep_regimes(r),
            "Le solde (ressources − dépenses) projeté, décomposé par groupe de régimes, hors "
            "transferts internes. Montre lesquels sont à l'équilibre et lesquels pèsent sur le "
            "déficit d'ensemble.",
            "COR, rapport 2026 (solde projeté par groupe de régimes, hors transferts internes, "
            "scénario de référence).")

    # --- Comparaisons & structures (axe catégoriel, barres)
    r = _rows("partie 3", "Fig 3.8")
    if r:
        cats, series = extract_stacked(
            r, 1.0,
            ["#6aa84f", "#1f4e79", "#e8731c", "#9c27b0", "#c2185b", "#b34700"])
        add("composition_revenu", "Composition du revenu disponible des ménages",
            "€/mois", " €", series,
            "D'où vient le revenu mensuel moyen des ménages, selon qu'ils sont retraités, "
            "actifs ou l'ensemble : revenus d'activité, pensions, patrimoine, prestations, "
            "moins les impôts et prélèvements sociaux (segments négatifs). Le repère = revenu "
            "disponible total (RdB).",
            "COR, rapport 2026 (composition du revenu disponible des ménages en 2023).",
            chartType="bar", barMode="stacked", categories=cats)

    r = _rows("partie 4", "Fig.4.C")
    if r:
        cats, series = extract_country_grouped(r, 1.0)
        add("emploi_pays", "Taux d'emploi par pays et tranche d'âge", "%", " %", series,
            "Où se situe la France ? Taux d'emploi par tranche d'âge (jeunes, seniors, âges "
            "intermédiaires) dans les pays suivis par le COR. La France se distingue par un "
            "emploi des seniors (60-64 ans) plus faible que la moyenne.",
            "COR, rapport 2026 (taux d'emploi par âge, comparaison internationale).",
            chartType="bar", barMode="grouped", categories=cats)

    r = _rows("partie 1", "Fig 1.8")
    if r:
        cats, series = extract_age_profile(r, 1.0)
        add("profil_age", "Taux d'activité par âge et sexe", "%", " %", series,
            "Le profil du taux d'activité selon l'âge, par sexe, à différentes dates "
            "(1975 → 2070). On lit la déformation des carrières : entrée plus tardive, "
            "et surtout activité bien plus forte après 55 ans dans les projections. "
            "Trait plein = femmes, tireté = hommes.",
            "COR, rapport 2026 (taux d'activité par âge en 1975, 2024, 2050 et 2070).",
            xLabel="Tranche d'âge", categories=cats)

    r = _rows("partie 2", "Tab 2.5")
    if r:
        cats, series = extract_determinants(r, "2070", 1.0)
        add("determinants_depenses", "Pourquoi la dépense projetée a changé (2070)",
            "points de PIB", " pts", series,
            "Décomposition de l'écart entre la dépense de retraite projetée en 2070 par le "
            "rapport 2026 et celle du rapport précédent : chaque facteur pousse à la hausse "
            "(vert) ou à la baisse (rouge), pour aboutir à l'écart total (bleu). La fécondité "
            "plus basse pèse le plus lourd. La barre « Autres / résidu » referme la "
            "décomposition (interactions et effets non détaillés).",
            "COR, rapport 2026 (Tableau 2.5, décomposition des écarts de dépenses en part de "
            "PIB, horizon 2070).",
            chartType="bar", waterfall=True, categories=cats)

    # --- Droits familiaux & réversion (rapport thématique COR, novembre 2025) ---
    DF = "2025-11"
    r = _rows("partie 1", "Fig 1.1", DF)
    if r:
        cats, series = extract_stacked(r, 100, DEFAULT_PALETTE)
        cats = [_short_regime(c) for c in cats]
        add("df_part_regime", "Poids des droits familiaux par régime", "%", " %", series,
            "Part des dispositifs de solidarité familiale (départs anticipés, majoration de "
            "pension, MDA, AVPF) dans les pensions de droit direct, par groupe de régimes. "
            "Ils pèsent près de 10 % au régime général.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.1).",
            chartType="bar", barMode="stacked", categories=cats)
    r = _rows("partie 1", "Fig 1.3", DF)
    if r:
        cats, series = extract_country_grouped(r, 1, ["#1f4e79", "#c2185b"])
        add("df_masses_genre", "Masses des droits familiaux versées, par sexe",
            "milliards d'euros", " Md€", series,
            "Montants annuels versés au titre de chaque dispositif familial, selon le sexe du "
            "bénéficiaire. La MDA et l'AVPF bénéficient très majoritairement aux femmes.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.3).",
            chartType="bar", barMode="grouped", categories=cats)
    r = _rows("partie 1", "Fig 1.6", DF)
    if r:
        series = extract_rows_timeseries(r, 100, 1990)
        obs = [s for s in series if "obs" in s["label"].lower()] or series
        series = obs[:1]
        if series:
            series[0].update(label="Part des réversions", color="#1f2d3d")
        add("df_reversion_part", "Part des pensions de réversion", "%", " %", series,
            "Part des pensions de réversion dans l'ensemble des pensions versées. La réversion "
            "reste un filet essentiel, surtout pour les femmes, mais sa part décline lentement.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.6).")
    r = _rows("partie 1", "Fig 1.17", DF)
    if r:
        series = extract_rows_timeseries(r, 1, 2000)
        add("df_beneficiaires", "Bénéficiaires de pensions, par type de droit",
            "milliers", " k", series,
            "Effectifs de retraités selon le type de droit perçu : droit direct seul, droits "
            "dérivés (réversion) seuls, ou les deux. Beaucoup de femmes cumulent droit direct "
            "et réversion.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.17).")
    r = _rows("partie 1", "Fig 1.8", DF)
    if r:
        add("df_age_enfants", "Âge de départ par sexe et nombre d'enfants", "ans", " ans",
            extract_cohort_groups(r, 1),
            "Âge moyen de départ à la retraite par génération, selon le sexe et le nombre "
            "d'enfants. Les mères de trois enfants et plus partent plus tôt (départs anticipés, "
            "MDA), les hommes partent globalement plus tôt que les femmes.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.8).",
            xLabel="Génération")
    r = _rows("partie 1", "Fig 1.9", DF)
    if r:
        add("df_duree_enfants", "Durée validée par sexe et nombre d'enfants", "trimestres",
            " trim.", extract_cohort_groups(r, 1),
            "Durée d'assurance validée moyenne par génération, selon le sexe et le nombre "
            "d'enfants. Les dispositifs familiaux (MDA, AVPF) compensent en partie les "
            "carrières plus courtes des mères.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.9).",
            xLabel="Génération")
    r = _rows("partie 1", "Fig 1.10", DF)
    if r:
        add("df_pension_enfants", "Pension relative par sexe et nombre d'enfants", "%", " %",
            extract_cohort_groups(r, 100),
            "Pension moyenne perçue à 68 ans, rapportée au salaire moyen (SMPT), selon le sexe "
            "et le nombre d'enfants. L'écart femmes-hommes et l'effet du nombre d'enfants "
            "restent marqués malgré les droits familiaux.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 1.10).",
            xLabel="Génération")
    r = _rows("partie 2", "Fig 2.1", DF)
    if r:
        series = extract_cols_timeseries(r, [(2, "Femmes", "#c2185b"), (3, "Hommes", "#1f4e79")], 1, 1960)
        add("df_inactivite", "Part des inactifs, par sexe (depuis 1968)", "%", " %", series,
            "Part des personnes inactives selon le sexe depuis 1968. L'inactivité féminine, "
            "très élevée autrefois, a fortement reculé — ce qui transforme peu à peu les droits "
            "à retraite des femmes.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 2.1).")
    r = _rows("partie 2", "Fig 2.7", DF)
    if r:
        series = extract_cols_timeseries(r, [(2, "Femmes", "#c2185b"), (3, "Hommes", "#1f4e79")], 1, 1970)
        add("df_emploi_genre", "Taux d'emploi des 15-64 ans, par sexe", "%", " %", series,
            "Taux d'emploi des femmes et des hommes (15-64 ans) depuis 1975 : la convergence "
            "est nette mais incomplète, et nourrit l'écart de pensions à venir.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 2.7).")
    r = _rows("partie 2", "Fig 2.9", DF)
    if r:
        series = extract_cols_timeseries(
            r, [(2, "Femmes", "#c2185b"), (3, "Hommes", "#1f4e79"), (4, "Ensemble", "#6aa84f")], 1, 1970)
        add("df_temps_partiel", "Part du temps partiel, par sexe", "%", " %", series,
            "Part de l'emploi à temps partiel selon le sexe. Le temps partiel, très féminin, "
            "réduit les salaires et donc les droits à retraite accumulés.",
            "COR, rapport « Droits familiaux et conjugaux » 2025 (Fig 2.9).")

    # --- Comparaisons internationales (Panorama des systèmes de retraite, 2020) ---
    PA = "2020-12"

    def intl_snap(iid, sheet, label, unit, suffix, scale, desc):
        r = _rows("thématique", sheet, PA)
        if not r:
            return
        cats, series = extract_category_snapshot(r, scale=scale)
        add(iid, label, unit, suffix, series,
            desc + " Comparaison internationale (instantané, données ~2017).",
            f"COR, « Panorama des systèmes de retraite en France et à l'étranger » 2020 ({sheet}).",
            chartType="bar", barMode="grouped", categories=cats)

    intl_snap("intl_emploi", "Fig 5.7", "Taux d'emploi des 20-64 ans, par pays", "%", " %", 100,
              "Part des 20-64 ans en emploi selon le pays.")
    intl_snap("intl_ratio_demo", "Fig 5.3", "Rapport démographique, par pays", "ratio", "", 1,
              "Personnes de 20-64 ans pour une personne de 65 ans et plus, selon le pays.")
    intl_snap("intl_revenus_act", "Fig 5.9", "Part des revenus d'activité dans le PIB, par pays",
              "%", " %", 100, "Poids de la rémunération du travail dans le PIB selon le pays.")
    intl_snap("intl_retraites", "Fig 5.11", "Taux de retraités, par pays", "%", " %", 100,
              "Nombre de retraités rapporté à la population de 65 ans et plus, selon le pays.")
    intl_snap("intl_pension_rel", "Fig 5.12", "Pension relative aux revenus d'activité, par pays",
              "%", " %", 100, "Pension moyenne rapportée au revenu d'activité moyen, selon le pays.")

    r = _rows("thématique", "Fig 5.1", PA)
    if r:
        cats, series = extract_stacked(r, 100, ["#1f4e79", "#8fb8e0"], total_keys=("totales", "total"))
        add("intl_depenses", "Dépenses de retraite (publiques + privées), par pays", "% du PIB",
            " %", series,
            "Part des dépenses de retraite dans le PIB selon le pays, distinguant les régimes "
            "publics des dispositifs privés (importants aux Pays-Bas, au Canada…). Comparaison "
            "internationale (données ~2017).",
            "COR, « Panorama des systèmes de retraite … » 2020 (Fig 5.1).",
            chartType="bar", barMode="stacked", categories=cats)
    r = _rows("thématique", "Fig 3.1", PA)
    if r:
        cats, series = extract_country_grouped(r, 100, ["#1f4e79", "#e8731c"])
        series = [s for s in series if "total" not in s["label"].lower()]
        add("intl_cotisation", "Taux de cotisation retraite, par pays", "%", " %", series,
            "Taux de cotisation aux régimes obligatoires selon le pays, part salariale et part "
            "employeur empilées. Comparaison internationale (données ~2017).",
            "COR, « Panorama des systèmes de retraite … » 2020 (Fig 3.1).",
            chartType="bar", barMode="stacked", categories=cats)

    # --- Projections par régime (rapport thématique COR, novembre 2017) ---
    RG = "2017-11"

    def reg_ind(iid, sheet, sub, label, unit, suffix, scale, desc):
        r = _rows("annexe", sheet, RG)
        if not r:
            return
        add(iid, label, unit, suffix, extract_regime_series(r, sub, scale),
            desc + " Une courbe par régime (projections du rapport thématique 2017).",
            f"COR, rapport « Perspectives … résultats par régime » 2017 ({sheet}).")

    reg_ind("reg_cotisants", "Tableau_1", "ensemble", "Cotisants par régime", "milliers", " k", 1,
            "Effectifs de cotisants projetés, par régime.")
    reg_ind("reg_retraites", "Tableau_2", "ensemble", "Retraités de droit direct par régime",
            "milliers", " k", 1, "Effectifs de retraités de droit direct projetés, par régime.")
    reg_ind("reg_ressources", "Tableau_4", "1,3", "Masse des ressources par régime", "% du PIB",
            " %", 100, "Ressources de chaque régime en part de PIB (scénario central 1,3 %).")
    reg_ind("reg_prestations", "Tableau_5", "1,3", "Masse des prestations par régime", "% du PIB",
            " %", 100, "Prestations de chaque régime en part de PIB (scénario central 1,3 %).")
    reg_ind("reg_depenses", "Tableau_6", "1,3", "Masse des dépenses par régime", "% du PIB",
            " %", 100, "Dépenses de chaque régime en part de PIB (scénario central 1,3 %).")
    reg_ind("reg_solde", "Tableau_7", "1,3", "Solde technique par régime", "% du PIB", " %", 100,
            "Solde technique (ressources − dépenses) de chaque régime, en part de PIB "
            "(scénario central 1,3 %).")
    reg_ind("reg_solde_elargi", "Tableau_8", "1,3", "Solde élargi par régime", "% du PIB", " %", 100,
            "Solde élargi (intégrant contributions et subventions d'équilibre) de chaque régime, "
            "en part de PIB (scénario central 1,3 %).")

    # --- Restes des rapports annuels : fonction publique (rapport 2026) ---
    r = _rows("partie 1", "Fig 1.13")
    if r:
        add("fp_effectifs", "Effectifs cotisants de la fonction publique", "base 100 en 2025",
            "", extract_rows_timeseries(r, 1, 2018, label_col=2),
            "Évolution des effectifs cotisants aux régimes de la fonction publique (base 100 "
            "en 2025), par versant. Le nombre de cotisants conditionne les ressources de ces "
            "régimes.",
            "COR, rapport 2026 (Fig 1.13).")
    r = _rows("partie 1", "Fig 1.14")
    if r:
        add("fp_remunerations", "Rémunérations moyennes dans la fonction publique",
            "base 100 en 2025", "", extract_rows_timeseries(r, 1, 2018, label_col=3),
            "Évolution des rémunérations moyennes des cotisants de la fonction publique (base "
            "100 en 2025) : traitement indiciaire, salaire moyen y compris primes. Elles "
            "déterminent l'assiette de cotisation et les pensions futures.",
            "COR, rapport 2026 (Fig 1.14).")
    r = _rows("partie 1", "Fig 1.15")
    if r:
        series = [s for s in extract_rows_timeseries(r, 100, 2018) if "vari" not in s["label"].lower()]
        add("fp_primes", "Part des primes dans la rémunération publique", "%", " %", series,
            "Part des primes (hors traitement indiciaire) dans la rémunération des "
            "fonctionnaires, par versant. Les primes ne cotisent que faiblement à la retraite, "
            "ce qui pèse sur les pensions de la fonction publique.",
            "COR, rapport 2026 (Fig 1.15).")

    # --- Sensibilité : faisceaux « et si l'hypothèse était différente ? »
    def dep_fan(rows):
        """Bloc 'Dépenses, en % du PIB' d'une figure de sensibilité : {clé: série}.

        Les sous-libellés sont en général 'Sc. Ref' / 'Var …' ; pour la
        sensibilité à la productivité (fig. 2.22 de 2026), ce sont des taux
        numériques (0.01, 0.004…) qu'on normalise en clés texte ('0.01').
        """
        ym = _ymap(rows)
        res, started = {}, False
        for r in rows:
            c1 = r[1] if len(r) > 1 else None
            c2 = r[2] if len(r) > 2 else None
            if c1 and str(c1).startswith("Dépenses, en % du PIB"):
                started = True
            if started and c1 and "Solde" in str(c1):
                break
            if not started:
                continue
            k = None
            if isinstance(c2, str) and (c2.strip() == "Sc. Ref" or c2.strip().startswith("Var")):
                k = c2.strip()
            elif isinstance(c2, float):
                k = format(c2, "g")
            if k:
                res[k] = {ym[i]: round(r[i] * 100, 3) for i in ym
                          if i < len(r) and isinstance(r[i], (int, float))}
        return res

    FANS = [
        ("sens_fec", "Si la fécondité changeait…", "Fig 2.18",
         [("fécondité haute", "Fécondité 1,7", "#2ca089"),
          ("fécondité basse", "Fécondité 1,2", "#d6452a")]),
        ("sens_ev", "Si on vivait plus ou moins longtemps…", "Fig 2.19",
         [("mortalité basse (EV haute)", "Espérance de vie haute", "#d6452a"),
          ("mortalité haute (EV basse)", "Espérance de vie basse", "#2ca089")]),
        ("sens_mig", "Si les migrations changeaient…", "Fig 2.20",
         [("smi haut", "Migrations hautes (+230 k)", "#2ca089"),
          ("smi bas", "Migrations basses (+70 k)", "#d6452a")]),
        ("sens_cho", "Si le chômage changeait…", "Fig 2.21",
         [("C5", "Chômage 5 %", "#2ca089"), ("C10", "Chômage 10 %", "#d6452a")]),
        ("sens_prod", "Si la productivité changeait…", "Fig 2.22",
         [("0.01", "Productivité 1,0 %", "#2ca089"),
          ("0.004", "Productivité 0,4 %", "#d6452a")]),
    ]
    for iid, label, sheet, variants in FANS:
        rows = _rows("partie 2", sheet)
        if not rows:
            continue
        fan = dep_fan(rows)
        ref = fan.get("Sc. Ref", {})
        if not ref:
            continue
        series = [{"label": "Scénario de référence", "color": "#1f4e79", "kind": "solid",
                   "points": [{"x": y, "y": ref[y]} for y in sorted(ref) if 2015 <= y <= 2070]}]
        spread = []
        for needle, vlabel, vcolor in variants:
            key = next((k for k in fan if needle.lower() in k.lower()), None)
            if not key:
                continue
            v = fan[key]
            series.append({"label": vlabel, "color": vcolor, "kind": "dash",
                           "points": [{"x": y, "y": v[y]} for y in sorted(v) if 2015 <= y <= 2070]})
            if 2070 in v:
                spread.append(v[2070])
        rng = (f" En 2070, les dépenses iraient de {min(spread):.1f} % à {max(spread):.1f} % du PIB "
               f"selon l'hypothèse (référence : {ref.get(2070, 0):.1f} %).") if len(spread) == 2 else ""
        b = _bounds(series)
        ind[iid] = {"label": label, "unit": "% du PIB", "suffix": " %",
                    "desc": "Dépenses de retraite en % du PIB selon l'hypothèse retenue." + rng,
                    "source": f"COR, rapport 2026 ({sheet}).", "series": series, **b}

    themes = [
        {"name": "Démographie", "indicators": ["cot_ret", "ratio_demo", "fecondite", "esp_vie", "migration"]},
        {"name": "Emploi & économie", "indicators": ["emploi", "chomage", "productivite", "pop_active", "emploi_seniors", "profil_age", "duree_carriere", "age_moyen_depart", "fp_effectifs", "fp_remunerations", "fp_primes"]},
        {"name": "Pensions & retraités", "indicators": ["age_depart", "pension_rel", "niveau_vie", "taux_rempl", "taux_cotisation", "tri", "duree_retraite", "ecart_genre", "assurance_genre", "age_depart_genre", "pauvrete", "intensite_pauvrete"]},
        {"name": "Finances du système", "indicators": ["depenses", "ressources", "solde", "solde_regimes", "dep_regimes", "dep_pub", "struct_ressources"]},
        {"name": "Comparaisons & structures", "indicators": ["composition_revenu", "emploi_pays", "determinants_depenses"]},
        {"name": "Droits familiaux & réversion", "indicators": ["df_part_regime", "df_masses_genre", "df_reversion_part", "df_beneficiaires", "df_age_enfants", "df_duree_enfants", "df_pension_enfants", "df_inactivite", "df_emploi_genre", "df_temps_partiel"]},
        {"name": "Comparaisons internationales", "indicators": ["intl_emploi", "intl_ratio_demo", "intl_revenus_act", "intl_retraites", "intl_pension_rel", "intl_depenses", "intl_cotisation"]},
        {"name": "Par régime (2017)", "indicators": ["reg_cotisants", "reg_retraites", "reg_ressources", "reg_prestations", "reg_depenses", "reg_solde", "reg_solde_elargi"]},
        {"name": "Sensibilité : et si… ?", "indicators": ["sens_fec", "sens_ev", "sens_mig", "sens_cho", "sens_prod"]},
    ]
    # on ne garde que les indicateurs réellement extraits
    for t in themes:
        t["indicators"] = [i for i in t["indicators"] if i in ind]
    themes = [t for t in themes if t["indicators"]]
    print(f"✓ explorateur : {len(ind)} indicateurs")
    return {"themes": themes, "indicators": ind}


def extract_international(path):
    """Part des dépenses de retraite (publiques/privées) dans le PIB par pays.

    Gère les deux mises en page du COR :
    - 2025 (« Part des dépenses OCDE ») : pays en ligne 4, années 2000/2021 en
      ligne 5, puis publiques/privées ;
    - 2026 (« Dépenses_OCDE ») : bloc « En 2021 » avec un pays par colonne,
      lignes « Publiques » / « Privées » (premier bloc = rapporté au PIB).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in ("Dépenses_OCDE", "Part des dépenses OCDE")
                  if s in wb.sheetnames), None)
    if sheet is None:
        wb.close()
        return None
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    countries, year = [], 2021

    # Mise en page 2026 : ligne « En AAAA | pays… » suivie de Publiques/Privées.
    for i, r in enumerate(rows[:10]):
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        if c1.startswith("En ") and c1[3:7].isdigit() and i + 2 < len(rows) \
                and str(rows[i + 1][1]).strip().startswith("Publiques"):
            year = int(c1[3:7])
            names, pub, priv = r, rows[i + 1], rows[i + 2]
            for j in range(2, len(names)):
                nm = names[j]
                if not (isinstance(nm, str) and nm.strip()):
                    continue
                p = pub[j] if j < len(pub) and isinstance(pub[j], (int, float)) else 0
                v = priv[j] if j < len(priv) and isinstance(priv[j], (int, float)) else 0
                countries.append({"name": nm.strip(),
                                  "pub": round(p * 100, 1), "priv": round(v * 100, 1),
                                  "total": round((p + v) * 100, 1)})
            break

    # Mise en page 2025 : pays et colonnes d'années côte à côte.
    if not countries:
        names, years, pub, priv = rows[3], rows[4], rows[5], rows[6]
        for i, nm in enumerate(names):
            if isinstance(nm, str) and nm.strip() and not nm.lower().startswith(("lecture", "champ", "source")):
                col = next((j for j in range(i, min(i + 5, len(years)))
                            if isinstance(years[j], int) and years[j] == year), None)
                if col is None:
                    continue
                p = pub[col] if col < len(pub) and isinstance(pub[col], (int, float)) else 0
                v = priv[col] if col < len(priv) and isinstance(priv[col], (int, float)) else 0
                countries.append({"name": nm.strip(),
                                  "pub": round(p * 100, 1), "priv": round(v * 100, 1),
                                  "total": round((p + v) * 100, 1)})
    if not countries:
        return None
    countries.sort(key=lambda c: c["total"], reverse=True)
    return {"year": year, "countries": countries}


def extract_leviers(path, sheet="Fig 2.24"):
    """Calibrage des 3 leviers : ajustement (via un seul levier) pour équilibrer en 2070."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb[sheet].iter_rows(values_only=True))
    ym = {i: c for i, c in enumerate(rows[3]) if isinstance(c, int) and 1990 <= c <= 2100}
    col = next((i for i, c in ym.items() if c == 2070), None)
    if col is None:
        wb.close()
        return None

    def g(i):
        return rows[i][col] if col < len(rows[i]) and isinstance(rows[i][col], (int, float)) else None

    pen_ref, pen_eq = g(4), g(5)          # pension relative
    age_ref, age_eq = g(6), g(7)          # âge effectif moyen
    tx_ref, tx_eq = g(8), g(9)            # taux de prélèvement
    wb.close()
    return {
        "horizon": 2070,
        "age": {"ref": round(age_ref, 2), "full_years": round(age_eq - age_ref, 2)},
        "cotis": {"ref": round(tx_ref * 100, 1), "full_pts": round((tx_eq - tx_ref) * 100, 2)},
        "pension": {"ref_pct": round(pen_ref * 100, 1),
                    "full_pct": round((pen_ref - pen_eq) / pen_ref * 100, 1)},
        "source": f"COR, rapport 2026 ({sheet.lower().replace('fig ', 'fig. ')}) — "
                  "niveau de chaque levier pour équilibrer en 2070.",
    }


def build():
    extracted = {}
    realised = None  # série observée la plus récente (rapport LATEST)
    for vy, dpat, fpat, prod in VINTAGES:
        path = first_file(dpat, fpat)
        if not path:
            print("✗ fichier introuvable :", vy)
            continue
        data = extract_depenses(path, prod)
        if not data or not data["projection"]:
            print("✗ bloc dépenses introuvable :", vy)
            continue
        extracted[vy] = data
        if vy == LATEST:
            realised = data["observed"]
        print(f"✓ {vy} : {len(data['projection'])} points de projection")

    # Série « réalisé » : on prend l'observé du rapport le plus récent.
    realised = realised or next(iter(extracted.values()))["observed"]
    realise_points = [{"x": y, "y": realised[y]} for y in sorted(realised)]

    projections = []
    for vy, _, _, _ in VINTAGES:
        if vy not in extracted:
            continue
        proj = extracted[vy]["projection"]
        years = sorted(proj)
        pts = [{"x": y, "y": proj[y]} for y in years]
        last = years[-1]
        projections.append({
            "label": f"Rapport {vy} (prod. {PROD_LABEL[vy]} %)",
            "year": int(vy),
            "color": COLORS[vy],
            "endNote": vy,
            "source": f"COR, rapport annuel {vy} — dépenses du système de retraite "
                      f"en % du PIB, scénario {PROD_LABEL[vy]} %.",
            "points": pts,
        })

    # ---- Solde et ressources : projections de TOUS les millésimes disponibles.
    #      Réalisé = série observée du rapport le plus récent.
    solde_multi = extract_fin_multi(SOLDE_SOURCES, "solde")
    ressources_multi = extract_fin_multi(RESSOURCES_SOURCES, "ressources")
    conv_by_vy = {vy: conv for vy, _, _, _, _, _, conv in SOLDE_SOURCES}

    sdr_latest = None
    latest_synth = first_file(f"{LATEST}-06", "synthèse")
    if latest_synth:
        sdr_latest = extract_sdr(latest_synth, "Solde dépenses ressources")
    solde_obs = (sdr_latest or {}).get("Solde", {})
    solde_realise = [{"x": y, "y": solde_obs[y]} for y in sorted(solde_obs) if y <= int(LATEST)]

    solde_proj = []
    for vy in sorted(solde_multi):
        solde = solde_multi[vy]
        year = int(vy)
        # projection = à partir de l'année du rapport ; réalisé = avant.
        pts = [{"x": y, "y": solde[y]} for y in sorted(solde) if y >= year]
        if not pts:
            continue
        solde_proj.append({
            "label": f"Rapport {vy} ({PROD_LABEL[vy]} %, {conv_by_vy[vy]})",
            "year": year, "color": COLORS[vy], "endNote": vy,
            "source": f"COR, rapport annuel {vy} — solde du système de retraite, "
                      f"scénario {PROD_LABEL[vy]} %, {conv_by_vy[vy]}.",
            "points": pts,
        })
        print(f"✓ solde {vy} : 2070={solde.get(2070)}  pts proj={len(pts)}")

    solde_block = None
    if solde_proj:
        ys = [p["y"] for pr in solde_proj for p in pr["points"]] \
            + [p["y"] for p in solde_realise]
        solde_block = {
            "title": "Le solde du système de retraite plonge dans les projections récentes",
            "subtitle": "Solde (ressources − dépenses) en % du PIB — scénario de référence de "
                        "chaque rapport (convention « COR » jusqu'en 2019, EPR ensuite)",
            "yLabel": "% du PIB",
            "yMin": round(min(ys) - 0.3, 1), "yMax": round(max(ys) + 0.3, 1),
            "xMin": 2000, "xMax": 2070,
            "realise": {"label": "Solde réalisé (observé)", "color": "#1f2d3d",
                        "kind": "solid", "points": solde_realise or []},
            "projections": solde_proj,
        }

    # ---- Dépenses vs Ressources (effet « ciseaux »), rapport le plus récent
    ciseaux_block = None
    if sdr_latest:
        dep = sdr_latest.get("Dépenses", {})
        res = sdr_latest.get("Ressources", {})
        ys = sorted(set(dep) | set(res))
        ciseaux_block = {
            "title": "Pourquoi un déficit ? Les dépenses montent, les ressources baissent",
            "subtitle": f"Système de retraite en % du PIB — scénario de référence du rapport {LATEST}",
            "yLabel": "% du PIB", "yMin": 12, "yMax": 15.5, "xMin": 2000, "xMax": 2070,
            "series": [
                {"label": "Dépenses", "color": "#c2185b", "kind": "solid",
                 "points": [{"x": y, "y": dep[y]} for y in ys if y in dep]},
                {"label": "Ressources", "color": "#2f6fb0", "kind": "solid",
                 "points": [{"x": y, "y": res[y]} for y in ys if y in res]},
            ],
        }

    # ---- Niveau de vie relatif des retraités : un millésime par rapport (2023+)
    nv_obs, nv_projs = {}, {}
    for vy, dpat, sheet, selector in NV_SOURCES:
        path = first_file(dpat, "synthèse")
        if not path:
            print("✗ niveau de vie : fichier introuvable", vy)
            continue
        o, ref = extract_nv_vintage(path, sheet, selector)
        if ref:
            nv_projs[vy] = ref
        else:
            print("✗ niveau de vie : projection introuvable", vy)
        if vy == LATEST and o:
            nv_obs = o

    niveau_block = None
    if nv_obs and nv_projs:
        # On démarre à 1996 (données annuelles continues) pour éviter les longues
        # interpolations des points épars d'avant.
        obs_pts = [{"x": y, "y": nv_obs[y]} for y in sorted(nv_obs) if y >= 1996]
        nv_proj_list = []
        for vy in sorted(nv_projs):
            ref = nv_projs[vy]
            pts = [{"x": y, "y": ref[y]} for y in sorted(ref) if y >= int(vy)]
            nv_proj_list.append({
                "label": f"Rapport {vy}", "year": int(vy), "color": COLORS[vy],
                "endNote": vy,
                "source": f"COR, rapport annuel {vy} — niveau de vie relatif des retraités, "
                          "scénario de référence.",
                "points": pts,
            })
        _nv_all = [{"points": obs_pts}] + [{"points": p["points"]} for p in nv_proj_list]
        _nv_b = _bounds(_nv_all)
        niveau_block = {
            "title": "Le niveau de vie des retraités décrocherait peu à peu",
            "subtitle": "Niveau de vie moyen des retraités rapporté à celui de l'ensemble de la "
                        "population (100 % = parité) — projection de chaque rapport",
            "yLabel": "%", "yMin": _nv_b["yMin"], "yMax": _nv_b["yMax"],
            "xMin": 1996, "xMax": 2070,
            "realise": {"label": "Observé", "color": "#1f2d3d", "kind": "solid", "points": obs_pts},
            "projections": nv_proj_list,
        }
        print(f"✓ niveau de vie : obs {len(obs_pts)} pts, {len(nv_proj_list)} millésimes "
              f"(2070 : {', '.join(f'{vy}={nv_projs[vy].get(2070)}' for vy in sorted(nv_projs))})")

    # ---- Âge conjoncturel : un millésime par rapport (2023+), pour l'explorateur
    age_obs, age_projs = {}, {}
    for vy, dpat, sheet in AGE_SOURCES:
        path = first_file(dpat, "synthèse")
        if not path:
            continue
        o, p = extract_age_vintage(path, sheet)
        if p:
            age_projs[vy] = {y: p[y] for y in p if y >= int(vy)}
        if vy == LATEST and o:
            age_obs = o
    if age_projs:
        print(f"✓ âge conjoncturel : {len(age_projs)} millésimes "
              f"(2070 : {', '.join(f'{vy}={age_projs[vy].get(2070)}' for vy in sorted(age_projs))})")

    # ---- Fécondité : observé (réel) + hypothèse centrale de trois époques
    fecondite_block = None
    f_2026 = extract_fecondite(first_file("2026-06", "partie 1") or "")
    f_2025 = extract_fecondite(first_file("2025-06", "juin 2025 - partie 1") or "")
    f_old = extract_fecondite(first_file("2019-06", "Partie 1") or "")
    if f_2026 and f_2026["observed"]:
        obs = f_2026["observed"]
        obs_pts = [{"x": y, "y": obs[y]} for y in sorted(obs)]

        def central_line(fec, ymin):
            c = fec["central"]
            return [{"x": y, "y": c[y]} for y in sorted(c) if y >= ymin]
        hyps = []
        if f_old and f_old["central"]:
            hyps.append({"label": "Hypothèse centrale 2016-2021 (1,95)",
                         "color": "#2ca089", "kind": "dash", "endNote": "1,95",
                         "points": central_line(f_old, 2017)})
        if f_2025 and f_2025["central"]:
            hyps.append({"label": "Hypothèse centrale 2022-2025 (1,80)",
                         "color": "#e8731c", "kind": "dash", "endNote": "1,80",
                         "points": central_line(f_2025, 2023)})
        if f_2026["central"]:
            hyps.append({"label": "Hypothèse centrale 2026 (1,45)",
                         "color": "#7b1fa2", "kind": "dash", "endNote": "1,45",
                         "points": central_line(f_2026, 2025)})
        fecondite_block = {
            "title": "Fécondité : l'hypothèse, longtemps trop haute, passe sous la réalité",
            "subtitle": "Indice conjoncturel de fécondité (enfants par femme)",
            "yLabel": "", "yMin": 1.35, "yMax": 2.05, "xMin": 2000, "xMax": 2050,
            "realise": {"label": "Fécondité réelle observée", "color": "#1f2d3d",
                        "kind": "solid", "points": [p for p in obs_pts if p["x"] >= 2000]},
            "hypotheses": hyps,
        }
        print(f"✓ fécondité : obs→{obs_pts[-1]['x']}={obs_pts[-1]['y']}  hypothèses={len(hyps)}")

    # ---- Productivité réelle (moyenne mobile) vs hypothèses
    prod_block = None
    annual = extract_productivite_obs(first_file("2026-06", "partie 1") or "")
    if annual:
        ma = [p for p in moving_average(annual, 5) if p["x"] >= 2000]
        xs = [p["x"] for p in ma]
        x0, x1 = min(xs), max(xs)
        prod_block = {
            "title": "Productivité : ce que le COR suppose vs ce qui se passe vraiment",
            "subtitle": "Croissance de la productivité du travail (%/an, moyenne mobile 5 ans pour l'observé)",
            "yLabel": "% / an", "yMin": -0.5, "yMax": 3, "xMin": 2000, "xMax": 2030,
            "realise": {"label": "Productivité réellement observée (moy. mobile)",
                        "color": "#1f2d3d", "kind": "solid", "points": ma},
            "hypotheses": [
                {"label": "Hypothèse 1,3 % (rapports jusqu'à 2022)", "color": "#d6452a",
                 "kind": "dash", "endNote": "1,3 %",
                 "points": [{"x": x0, "y": 1.3}, {"x": 2030, "y": 1.3}]},
                {"label": "Hypothèse 1,0 % (rapports 2023-2024)", "color": "#e8731c",
                 "kind": "dash", "endNote": "1,0 %",
                 "points": [{"x": x0, "y": 1.0}, {"x": 2030, "y": 1.0}]},
                {"label": "Hypothèse 0,7 % (référence depuis 2025)", "color": "#c2185b",
                 "kind": "dash", "endNote": "0,7 %",
                 "points": [{"x": x0, "y": 0.7}, {"x": 2030, "y": 0.7}]},
            ],
        }
        print(f"✓ productivité obs : {x0}-{x1}, {len(ma)} pts")

    # ---- Hypothèses de la partie 1, sondées sur TOUS les millésimes : on
    #      tente l'extraction partout, on ne garde que ce qui existe (les
    #      formats trop anciens — 2016 et avant — sont écartés d'eux-mêmes).
    ALL_P1 = [("2017", "2017-06", "Contexte"), ("2018", "2018-06", "contexte"),
              ("2019", "2019-06", "partie 1"), ("2020", "2020-11", "partie 1"),
              ("2021", "2021-06", "partie 1"), ("2022", "2022-09", "septembre 2022 - partie 1"),
              ("2023", "2023-06", "partie 1"), ("2024", "2024-06", "partie 1"),
              ("2025", "2025-06", "juin 2025 - partie 1"), ("2026", "2026-06", "partie 1")]

    ev_hyps, mig_hyps, fec_hyps, ratio_hyps = {}, {}, {}, {}
    for vy, dpat, fpat in ALL_P1:
        path = first_file(dpat, fpat)
        if not path:
            continue
        s = labeled_row(path, ("espérance de vie", "65 ans"), "scénario central")
        if s:
            ev_hyps[vy] = {y: v for y, v in s.items() if y >= int(vy)}
        s = labeled_row(path, ("solde migratoire",), "scénario central", 0.001)
        if s:
            mig_hyps[vy] = {y: v for y, v in s.items() if y >= int(vy)}
        f = extract_fecondite(path)
        if f and f.get("central"):
            fec_hyps[vy] = {y: v for y, v in f["central"].items() if y >= int(vy)}
        s = block_labeled_row(path, ("rapport", "démographique"),
                              "20-64 / 65+", "scénario central")
        if s:
            ratio_hyps[vy] = {y: v for y, v in s.items() if y >= int(vy)}
    print(f"✓ espérance de vie : {len(ev_hyps)} millésimes "
          f"(2070 : {', '.join(f'{vy}={ev_hyps[vy].get(2070)}' for vy in sorted(ev_hyps))})")
    print(f"✓ migration : {len(mig_hyps)} millésimes")
    print(f"✓ fécondité (hypothèses) : {len(fec_hyps)} millésimes "
          f"(2070 : {', '.join(f'{vy}={fec_hyps[vy].get(2070)}' for vy in sorted(fec_hyps))})")
    print(f"✓ rapport démographique : {len(ratio_hyps)} millésimes")

    # ---- Chômage, taux d'emploi, productivité : le scénario retenu par
    #      chaque rapport (le libellé porte l'hypothèse : 7 %, 4,5 %, 1,3 %…
    #      → annotation de légende)
    REF_KEYS = ("tous scénarios", "scénario de référence")
    PROD_KEYS = {"1,3": ("scénario de référence", "scénario 1,3"),
                 "1,0": ("scénario de référence", "scénario 1,0"),
                 "0,7": ("scénario de référence", "scénario 0,7")}
    cho_hyps, cho_notes = {}, {}
    emp_hyps, emp_notes = {}, {}
    prod_hyps, prod_notes = {}, {}
    for vy, dpat, fpat in ALL_P1:
        path = first_file(dpat, fpat)
        if not path:
            continue
        s, lbl = labeled_row_any(path, ("chômage", "projeté"), REF_KEYS, 100)
        if s:
            cho_hyps[vy] = {y: v for y, v in _pct_fix(s).items() if y >= int(vy)}
            cho_notes[vy] = _hyp_note(lbl)
        s, lbl = labeled_row_any(path, ("taux d'emploi", "projeté"), REF_KEYS, 100)
        if s:
            emp_hyps[vy] = {y: v for y, v in _pct_fix(s).items() if y >= int(vy)}
            emp_notes[vy] = _hyp_note(lbl)
        s, lbl = labeled_row_any(path, ("productivité", "observés"),
                                 PROD_KEYS.get(PROD_LABEL.get(vy, ""), REF_KEYS), 100)
        if s:
            prod_hyps[vy] = {y: v for y, v in s.items() if y >= int(vy)}
            prod_notes[vy] = _hyp_note(lbl)
    print(f"✓ chômage : {len(cho_hyps)} millésimes "
          f"({', '.join(f'{vy}:{cho_notes[vy]}' for vy in sorted(cho_hyps))})")
    print(f"✓ emploi : {len(emp_hyps)} millésimes")
    print(f"✓ productivité projetée : {len(prod_hyps)} millésimes "
          f"({', '.join(f'{vy}:{prod_notes[vy]}' for vy in sorted(prod_hyps))})")

    # ---- Pension moyenne / salaire moyen : scénario de référence par rapport
    #      (avant 2022, la figure n'isole pas ces deux blocs : hors champ)
    pension_projs = {}
    for vy, dpat, fpat, ref in [
            ("2022", "2022-09", "septembre 2022 - partie 2", 0.013),
            ("2023", "2023-06", "partie 2", 0.010),
            ("2024", "2024-06", "partie 2", "Sc. Ref"),
            ("2025", "2025-06", "juin 2025 - partie 2", "Sc. Ref"),
            ("2026", "2026-06", "partie 2", "Sc. Ref")]:
        s = pension_ratio_vintage(first_file(dpat, fpat), ref)
        if s:
            pension_projs[vy] = {y: v for y, v in s.items() if y >= int(vy)}
    print(f"✓ pension relative : {len(pension_projs)} millésimes "
          f"(2070 : {', '.join(f'{vy}={pension_projs[vy].get(2070)}' for vy in sorted(pension_projs))})")

    # ---- Cotisants par retraité : données complémentaires 2024 et 2025
    cotret_projs = {}
    for vy, dpat in [("2024", "2024-06"), ("2025", "2025-06")]:
        cr = _rows("complémentaires", "Cotisants_Retraités", dpat)
        if not cr:
            continue
        cp = _block_sub(cr, "cotisants", "Sc. Ref")
        rp = _block_sub(cr, "retraités", "Sc. Ref")
        s = _ratio(cp, rp)
        if s:
            cotret_projs[vy] = {y: round(v, 3) for y, v in s.items() if y >= int(vy)}
    print(f"✓ cotisants/retraité : {len(cotret_projs)} millésimes")

    # ---- Taux de croissance de la population active : scénario de référence
    #      par rapport (la figure isole un taux observé puis projeté à partir
    #      de 2023 ; avant, les rapports ne publient que des moyennes de période)
    popact_projs = {}
    for vy, dpat in [("2023", "2023-06"), ("2024", "2024-06"),
                     ("2025", "2025-06"), ("2026", "2026-06")]:
        _o, p = extract_pop_active(first_file(dpat, "partie 1"), vy)
        if p:
            popact_projs[vy] = {y: v for y, v in p.items() if y >= int(vy)}
    print(f"✓ population active : {len(popact_projs)} millésimes")

    # ---- Parties 3 & 4 : séries par génération / par année, superposées par
    #      rapport. Repérage par titre (le numéro de figure change d'une édition
    #      à l'autre) ; garde-fou de plausibilité pour écarter les figures
    #      homonymes des vieux rapports (ex. « écart d'âge » vs « âge moyen »).
    ALL_VINT = [("2016", "2016-06"), ("2017", "2017-06"), ("2018", "2018-06"),
                ("2019", "2019-06"), ("2020", "2020-11"), ("2021", "2021-06"),
                ("2022", "2022-09"), ("2023", "2023-06"), ("2024", "2024-06"),
                ("2025", "2025-06"), ("2026", "2026-06")]

    def _prodfrac(vy):  # productivité de réf. → fraction (sélection du scénario)
        return {"1,3": 0.013, "1,0": 0.010, "0,7": 0.007}.get(PROD_LABEL.get(vy, ""), 0.013)

    taux_rempl_projs, duree_ret_projs, duree_car_projs = {}, {}, {}
    age_dep_projs, pauvrete_projs = {}, {}
    tri_projs, taux_cot_projs, intensite_projs = {}, {}, {}
    for vy, dpat in ALL_VINT:
        keys = ("taux de remplacement", "non-cadre")
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("sc. réf", "sc. ref", "scénario de référence", "référence",
                             PROD_LABEL.get(vy, "")), 100, 1900, _prodfrac(vy))
        s = _plausible(s, 30, 100)
        if s:
            taux_rempl_projs[vy] = s

        keys = ("durée de retraite",)
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("scénario central", "central"), 1, 1900,
                            block_key="moyenne par génération")
        s = _plausible(s, 15, 35)
        if s:
            duree_ret_projs[vy] = s

        keys = ("durée de carrière",)
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("moyenne par génération",), 1, 1900)
        s = _plausible(s, 25, 55)
        if s:
            duree_car_projs[vy] = s

        keys, exc = ("âge moyen de départ", "génération"), ("femmes",)  # ≠ figure F/H
        path = file_with_sheet(dpat, keys, exc)
        obs = pick_cohort_row(path, keys, ("âge moyen observé", "observé"), 1, 1900, exclude=exc)
        proj = pick_cohort_row(path, keys, ("âge moyen projeté", "projeté"), 1, 1900, exclude=exc)
        merged = _plausible({**proj, **obs}, 55, 70)
        if merged:
            age_dep_projs[vy] = merged

        keys = ("pauvreté", "retraité")
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("ensemble des retraités",), 100, 1990)
        s = _plausible(s, 3, 20)
        if s:
            pauvrete_projs[vy] = s

        keys = ("taux de rendement interne", "non-cadre")
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("sc. réf", "sc. ref", "scénario de référence", "référence"),
                            100, 1900)
        s = _plausible(s, 0, 10)
        if s:
            tri_projs[vy] = s

        keys = ("taux de cotisation", "retraite")
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys, ("tous régimes",), 100, 1900)
        s = _plausible(s, 10, 40)
        if s:
            taux_cot_projs[vy] = s

        keys = ("intensité de la pauvreté",)
        s = pick_cohort_row(file_with_sheet(dpat, keys), keys,
                            ("ensemble des retraités",), 100, 1990)
        s = _plausible(s, 5, 25)
        if s:
            intensite_projs[vy] = s

    print(f"✓ taux de remplacement : {len(taux_rempl_projs)} millésimes")
    print(f"✓ durée de retraite : {len(duree_ret_projs)} millésimes")
    print(f"✓ durée de carrière : {len(duree_car_projs)} millésimes")
    print(f"✓ âge moyen de départ : {len(age_dep_projs)} millésimes")
    print(f"✓ pauvreté retraités : {len(pauvrete_projs)} millésimes")
    print(f"✓ TRI cas-type : {len(tri_projs)} millésimes")
    print(f"✓ taux de cotisation : {len(taux_cot_projs)} millésimes")
    print(f"✓ intensité pauvreté : {len(intensite_projs)} millésimes")

    # Séries multi-millésimes mises à disposition de l'explorateur : pour
    # chaque indicateur, les projections de tous les rapports qui publient
    # le même jeu de données, à superposer au réalisé.
    multi = {
        "depenses": {vy: d["projection"] for vy, d in extracted.items()},
        "ressources": ressources_multi,
        "solde": solde_multi,
        "nv": nv_projs,
        "age": age_projs,
        "esp_vie": ev_hyps,
        "migration": mig_hyps,
        "fecondite": fec_hyps,
        "chomage": cho_hyps,
        "emploi": emp_hyps,
        "prod_path": prod_hyps,
        "ratio_demo": ratio_hyps,
        "pension_rel": pension_projs,
        "cot_ret": cotret_projs,
        "pop_active": popact_projs,
        "taux_rempl": taux_rempl_projs,
        "duree_retraite": duree_ret_projs,
        "duree_carriere": duree_car_projs,
        "age_moyen_depart": age_dep_projs,
        "pauvrete": pauvrete_projs,
        "tri": tri_projs,
        "taux_cotisation": taux_cot_projs,
        "intensite_pauvrete": intensite_projs,
        "_notes": {"chomage": cho_notes, "emploi": emp_notes, "prod_path": prod_notes},
    }

    out = {
        "depensesPib": {
            "title": "La part des retraites dans le PIB selon les rapports successifs du COR",
            "subtitle": "Dépenses de retraite en % du PIB — scénario de référence de chaque rapport (données officielles du COR)",
            "yLabel": "% du PIB",
            "yMin": 11.5,
            "yMax": 15.5,
            "xMin": 2000,
            "xMax": 2070,
            "realise": {
                "label": "Réalisé (observé)",
                "color": "#1f2d3d",
                "kind": "solid",
                "points": realise_points,
            },
            "projections": projections,
        },
        "solde": solde_block,
        "ressourcesVsDepenses": ciseaux_block,
        "niveauVie": niveau_block,
        "fecondite": fecondite_block,
        "productiviteReel": prod_block,
        "international": extract_international(first_file(R26, "synthèse") or ""),
        "leviers": extract_leviers(first_file(R26, "partie 2") or ""),
    }
    # L'explorateur (de loin le plus gros bloc) part dans un fichier séparé,
    # chargé paresseusement côté client à l'approche de la section : le
    # chargement initial (téléchargement + JSON.parse) en est d'autant allégé.
    explorer = {"explorer": build_explorer(multi)}

    data_dir = os.path.join(os.path.dirname(__file__), "..", "data")
    series_dest = os.path.join(data_dir, "cor-series.generated.js")
    explorer_dest = os.path.join(data_dir, "cor-explorer.generated.js")
    write_js(series_dest, "COR_SERIES", out)
    write_js(explorer_dest, "COR_EXPLORER", explorer)
    print("\nÉcrit :", os.path.relpath(series_dest), "+", os.path.relpath(explorer_dest))
    print("Millésimes :", ", ".join(k for k in extracted))


if __name__ == "__main__":
    build()
